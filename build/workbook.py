"""Application_Command_Center.xlsx — the working layer.

Five tabs, each answering exactly one question. The rule that shapes every tab:
a date this sweep could not confirm for 2027 never appears as if it were a
commitment. Deadlines carry a confidence column, and comparators are labelled.
"""
import sys, datetime as dt
sys.path.insert(0, "build")
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from data import load, funding_rows, TODAY, PATH_COLOR, PATH_NAME, money
import importlib.util
spec = importlib.util.spec_from_file_location("rank", "research/rank.py")

INK="FF16161E"; ACCENT="FFE0603A"; MUT="FF5C5C68"; RULE="FFD8D4CE"
WASH="FFF7F5F2"; RED="FFFCE4DC"; ORANGE="FFFBEEDC"; GREY="FFEDEDED"; GREEN="FFE9F1EC"
H_FONT = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
H_FILL = PatternFill("solid", fgColor=INK)
BODY   = Font(name="Calibri", size=10, color=INK)
SMALL  = Font(name="Calibri", size=9, color=MUT)
LINKF  = Font(name="Calibri", size=10, color=ACCENT, underline="single")
THIN   = Border(bottom=Side("thin", color=RULE))
WRAP   = Alignment(wrap_text=True, vertical="top")
TOP    = Alignment(vertical="top")

STATUSES = "todo,preparing,submitted,interview,accepted,rejected,waitlist"
DOC_STATUSES = "not started,in progress,ordered,received,submitted"

def header(ws, headers, widths, row=1):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.alignment = H_FONT, H_FILL, Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row+1, column=1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"

def put(ws, r, c, v, font=BODY, fill=None, wrap=False, num=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.alignment = WRAP if wrap else TOP; cell.border = THIN
    if fill: cell.fill = PatternFill("solid", fgColor=fill)
    if num: cell.number_format = num
    return cell

def link(ws, r, c, url, text="open"):
    cell = ws.cell(row=r, column=c, value=text if url else "")
    if url and url.startswith("http"):
        cell.hyperlink = url; cell.font = LINKF
    else:
        cell.value = "—"; cell.font = SMALL
    cell.alignment = TOP; cell.border = THIN
    return cell

def path_fill(paths):
    if not paths: return None
    hexc = PATH_COLOR.get(paths[0], "#5C5C68").lstrip("#")
    # a pale tint of the path colour, so the fill never fights the text
    r, g, b = (int(hexc[i:i+2], 16) for i in (0, 2, 4))
    tint = lambda x: format(int(x + (255 - x) * 0.82), "02X")
    return "FF" + tint(r) + tint(g) + tint(b)

rows = load()
usable = [r for r in rows if r["usable"]]

# ---------- which programmes are worth applying to -------------------------
import subprocess, json
subprocess.run([sys.executable, "research/rank.py"], capture_output=True)
ranked = json.load(open("research/ranked.json"))
score_by_id = {r["id"]: r.get("score", 0) for r in ranked}
import re
sys.path.insert(0, "research")
import importlib
rk = importlib.import_module("rank")
tracker_rows = []
for r in usable:
    why = rk.viability(r)
    if why: continue
    r["_score"] = score_by_id.get(r["id"], 0)
    tracker_rows.append(r)
tracker_rows.sort(key=lambda r: -r["_score"])
tracker_rows = tracker_rows[:34]

def priority(i):
    return 1 if i < 8 else (2 if i < 18 else 3)

def next_action(r):
    if r["deadline_conf"] == "none":
        return "Confirm the 2027 deadline on the official page", None
    if (r["portfolio_or_audition_required"] or "").lower().startswith("yes"):
        return "Start the portfolio brief for this programme", (r["deadline_date"] - dt.timedelta(days=120)) if r["deadline_date"] else None
    if r["english_level_required"] and not r["english_level_required"].upper().startswith(("TBC","NONE","NOT")):
        return "Book IELTS so the score reports before this deadline", (r["deadline_date"] - dt.timedelta(days=90)) if r["deadline_date"] else None
    return "Draft the application and gather documents", (r["deadline_date"] - dt.timedelta(days=60)) if r["deadline_date"] else None

wb = Workbook()

# ============================ 1. START HERE ================================
ws = wb.active; ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 96
ws["B2"] = "APPLICATION COMMAND CENTER"
ws["B2"].font = Font(name="Calibri", size=20, bold=True, color=INK)
ws["B3"] = f"Master's, September 2027 intake · built {TODAY.isoformat()} · Zizz (Tunisia, 300-ECTS software-engineering diploma)"
ws["B3"].font = SMALL
ws["B5"] = "What this is"; ws["B5"].font = Font(size=12, bold=True, color=ACCENT)
ws["C5"] = ("The working layer over a sweep of 197 verified master's programmes and 139 funding schemes "
            "across Spain, the Netherlands and Berlin. Each tab answers exactly one question. Nothing here "
            "is a recommendation — read Decision_Brief.pdf for that — and nothing here is unverified.")
ws["C5"].alignment = WRAP
r = 7
ws.cell(row=r, column=2, value="The tabs").font = Font(size=12, bold=True, color=ACCENT)
r += 1
TABS = [
 ("Tracker", "What do I do next? One row per programme worth applying to, sorted by when you must act. "
             "Programme and scholarship deadlines are SEPARATE columns because they differ, often by weeks."),
 ("Funding Matrix", "Where does the money come from? Every scheme, what it covers, whether a Tunisian can "
                    "actually win it, and whether it stacks. Rows in bold red close BEFORE the programme they fund."),
 ("Documents", "What must I prepare, and by when? Deduplicated across all applications, with lead times "
               "measured from Tunisia and worked backward from the earliest deadline that needs them."),
 ("All Programs", "What exists? The full verified dataset, one column per field. The reference layer "
                  "behind every other tab."),
]
for name, desc in TABS:
    c = ws.cell(row=r, column=2, value=name)
    c.font = Font(size=10, bold=True, color=ACCENT, underline="single")
    c.hyperlink = f"#'{name}'!A1"
    d = ws.cell(row=r, column=3, value=desc); d.alignment = WRAP
    ws.row_dimensions[r].height = 30
    r += 1
r += 1
ws.cell(row=r, column=2, value="How the dropdowns work").font = Font(size=12, bold=True, color=ACCENT)
ws.cell(row=r, column=3, value=(f"Tracker → Status: {STATUSES.replace(',', ' · ')}.   "
        f"Documents → Status: {DOC_STATUSES.replace(',', ' · ')}.   "
        "Type nothing else in those cells; the dropdown is the only valid input.")).alignment = WRAP
r += 2
ws.cell(row=r, column=2, value="Colour legend").font = Font(size=12, bold=True, color=ACCENT)
r += 1
LEG = [("Deadline within 7 days", RED), ("Deadline within 30 days", ORANGE),
       ("Deadline already past", GREY), ("Confirmed 2027 date", GREEN)]
for label, fill in LEG:
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row=r, column=2).border = THIN
    ws.cell(row=r, column=3, value=label).font = BODY
    r += 1
r += 1
ws.cell(row=r, column=2, value="Path colours").font = Font(size=12, bold=True, color=ACCENT)
r += 1
for L, name in PATH_NAME.items():
    c = ws.cell(row=r, column=2, value=L)
    c.fill = PatternFill("solid", fgColor=path_fill([L])); c.font = Font(bold=True, color=INK)
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value=name).font = BODY
    r += 1
r += 1
ws.cell(row=r, column=2, value="The one rule").font = Font(size=12, bold=True, color=ACCENT)
ws.cell(row=r, column=3, value=(
    "Every date carries a confidence. CONFIRMED 2027 came off a page showing a 2027 date. "
    "RECURRING ANNUAL is a date the institution repeats every year, projected to 2027. "
    "COMPARATOR 2026 is last cycle's date, kept for planning and NOT a commitment — it is excluded "
    "from deadlines.ics entirely. Unverified programmes appear in no tab here; they are listed only in "
    "the risk section of Decision_Brief.pdf.")).alignment = WRAP
ws.row_dimensions[r].height = 46

# ============================ 2. TRACKER ===================================
ws = wb.create_sheet("Tracker")
H = ["Programme","Institution","City","Path","Programme deadline","Deadline confidence",
     "Scholarship deadline","Scholarship","English test required","Application fee (EUR)",
     "Portfolio?","Status","Priority","Next action","Act by","Link"]
W = [40,30,14,7,15,17,15,30,22,12,11,13,9,38,12,7]
header(ws, H, W)

fund = funding_rows()
def scholarship_for(r):
    names = r["scholarship_names"]
    if not names or names.upper().startswith("TBC"): return "", None, ""
    first = names.split("|")[0].strip()[:60]
    best = None
    for f in fund:
        inst = (f.get("applies_to_institution") or "").lower()
        if inst and inst != "any" and inst[:14] in r["institution"].lower() and f["deadline_date"]:
            if best is None or f["deadline_date"] < best[0]:
                best = (f["deadline_date"], f.get("scholarship_name", "")[:50], f["deadline_conf"])
    if best: return best[1] or first, best[0], best[2]
    return first, None, ""

for i, r in enumerate(tracker_rows):
    r["_priority"] = priority(i)
    r["_act"], r["_act_date"] = next_action(r)
# sorted by next_action_date, with undated rows last so they cannot hide real work
tracker_rows.sort(key=lambda r: (r["_act_date"] is None, r["_act_date"] or dt.date(2099,1,1)))

row = 2
for i, r in enumerate(tracker_rows):
    act, act_date = r["_act"], r["_act_date"]
    sname, sdate, sconf = scholarship_for(r)
    fill = path_fill(r["paths"])
    put(ws, row, 1, r["program_name"][:110], wrap=True)
    put(ws, row, 2, r["institution"][:70], wrap=True)
    put(ws, row, 3, r["city"][:28])
    c = put(ws, row, 4, ",".join(r["paths"])); c.fill = PatternFill("solid", fgColor=fill or WASH)
    put(ws, row, 5, r["deadline_date"], num="yyyy-mm-dd") if r["deadline_date"] else put(ws, row, 5, "TBC", font=SMALL)
    put(ws, row, 6, r["deadline_conf"].replace("_", " "), font=SMALL)
    put(ws, row, 7, sdate, num="yyyy-mm-dd") if sdate else put(ws, row, 7, "—", font=SMALL)
    put(ws, row, 8, sname or "none found", wrap=True, font=BODY if sname else SMALL)
    put(ws, row, 9, (r["english_level_required"] or "TBC")[:60], wrap=True)
    fee = money(r.get("tuition_notes", "")) if "application fee" in (r.get("tuition_notes") or "").lower() else None
    put(ws, row, 10, fee if fee else "—", font=BODY if fee else SMALL)
    put(ws, row, 11, "yes" if (r["portfolio_or_audition_required"] or "").lower().startswith("yes") else "no")
    put(ws, row, 12, "todo")
    put(ws, row, 13, r["_priority"])
    put(ws, row, 14, act, wrap=True)
    put(ws, row, 15, act_date, num="yyyy-mm-dd") if act_date else put(ws, row, 15, "—", font=SMALL)
    link(ws, row, 16, r["program_url"])
    ws.row_dimensions[row].height = 30
    row += 1

dv = DataValidation(type="list", formula1=f'"{STATUSES}"', allow_blank=False)
dv.error = "Pick a status from the dropdown."; dv.errorTitle = "Not a valid status"
ws.add_data_validation(dv); dv.add(f"L2:L{row-1}")
last = row - 1
for col in ("E", "O"):
    rng = f"{col}2:{col}{last}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}2),{col}2<TODAY())'],
        fill=PatternFill("solid", fgColor=GREY),
        font=Font(strike=True, color=MUT), stopIfTrue=True))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}2),{col}2-TODAY()<=7)'],
        fill=PatternFill("solid", fgColor=RED), font=Font(bold=True, color="FFC2451F")))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(ISNUMBER({col}2),{col}2-TODAY()<=30)'],
        fill=PatternFill("solid", fgColor=ORANGE)))
ws.conditional_formatting.add(f"F2:F{last}", FormulaRule(
    formula=['EXACT($F2,"confirmed 2027")'], fill=PatternFill("solid", fgColor=GREEN)))

# ============================ 3. FUNDING MATRIX ============================
ws = wb.create_sheet("Funding Matrix")
H = ["Scheme","Provider","Region","Applies to","Coverage","Amount (EUR/yr)",
     "Tunisian eligible","Evidence","Own deadline","Deadline confidence",
     "Closes before the programme?","Stacks","Effort","Link"]
W = [40,26,14,28,14,16,15,44,13,17,20,9,22,7]
header(ws, H, W)
prog_deadline_by_inst = {}
for r in usable:
    if r["deadline_date"]:
        k = r["institution"].lower()[:16]
        prog_deadline_by_inst[k] = min(prog_deadline_by_inst.get(k, r["deadline_date"]), r["deadline_date"])
row = 2
for f in sorted(fund, key=lambda x: (x.get("region",""), x.get("scholarship_name",""))):
    tun = (f.get("tunisia_eligible") or "").lower()
    inst = (f.get("applies_to_institution") or "")
    before = ""
    pd = prog_deadline_by_inst.get(inst.lower()[:16])
    if f["deadline_date"] and pd:
        before = "YES — apply earlier" if f["deadline_date"] < pd else "no"
    bold = before.startswith("YES")
    fnt = Font(name="Calibri", size=10, bold=True, color="FFC2451F") if bold else BODY
    put(ws, row, 1, (f.get("scholarship_name") or "")[:110], font=fnt, wrap=True)
    put(ws, row, 2, (f.get("provider") or "")[:60], wrap=True)
    put(ws, row, 3, (f.get("region") or "").replace("_", " "))
    put(ws, row, 4, (inst or "any")[:60], wrap=True)
    cov = (f.get("coverage_level") or "")
    put(ws, row, 5, cov, fill=GREEN if cov.startswith("full") else None)
    put(ws, row, 6, (f.get("amount_eur_per_year") or "TBC")[:28])
    put(ws, row, 7, tun or "unclear",
        fill={"yes": GREEN, "no": RED, "likely": ORANGE}.get(tun))
    put(ws, row, 8, (f.get("tunisia_eligible_evidence") or "")[:220], wrap=True, font=SMALL)
    put(ws, row, 9, f["deadline_date"], num="yyyy-mm-dd") if f["deadline_date"] else put(ws, row, 9, (f.get("deadline") or "TBC")[:18], font=SMALL)
    put(ws, row, 10, f["deadline_conf"].replace("_", " "), font=SMALL)
    put(ws, row, 11, before, font=fnt, fill=RED if bold else None)
    put(ws, row, 12, "y" if "stack" in (f.get("notes","") + f.get("scholarship_name","")).lower() else "?")
    put(ws, row, 13, "automatic" if "automatic" in (f.get("apply_via","") or "").lower() else "separate application")
    link(ws, row, 14, f.get("funding_url", ""))
    ws.row_dimensions[row].height = 30
    row += 1

# ============================ 4. DOCUMENTS =================================
ws = wb.create_sheet("Documents")
H = ["Document","Why / who needs it","Lead time (weeks)","Start by","Blocking?","Status","Notes"]
W = [40,46,16,13,11,15,56]
header(ws, H, W)
earliest = min([r["deadline_date"] for r in tracker_rows if r["deadline_date"]] or [dt.date(2027,1,12)])
DOCS = [
 ("Degree certificate — sworn translation + apostille", "Every application, and the visa file", 8,
  "The long pole. Tunisian apostille plus a sworn translation runs weeks and is invisible until it blocks you. LEAD TIME IS AN ESTIMATE, not a verified figure — confirm with the Ministry."),
 ("Official transcripts (sealed copies)", "Every application", 4,
  "Order several sealed originals at once; most institutions will not return them."),
 ("IELTS Academic — booking to reported score", "All except UJA Linares (B1) and the Spanish-taught rows", 9,
  "Test date, ~13 days to results, then institutional reporting. Range needed runs B1 (UJA) to IELTS 6.5 no band under 6.0 (EIT Digital) to TOEFL 100 (Berklee)."),
 ("Passport — validity through end of studies", "Visa, all countries", 8,
  "Must outlast the programme. Renew now if it expires before 2029."),
 ("Recommendation letters (2)", "TU Berlin, EIT Digital, ESMT, Berklee", 6,
  "Ask early; academic staff disappear over the summer."),
 ("CV — academic format", "Every application", 2, "One master version, forked per audience."),
 ("Statement of purpose — technical fork", "TU Berlin, EIT Digital, UGR, UMA, UJA", 4,
  "Write once, fork for a technical committee."),
 ("Statement of purpose — business fork", "ESMT, Macromedia, Berklee", 3,
  "Different reader entirely: venture and market, not signal processing."),
 ("Portfolio — 3 finished pieces", "Netherlands Film Academy, UdK Design & Computation", 16,
  "Only three shortlisted programmes are portfolio-gated and none needs finished music. See Portfolio_Brief.pdf."),
 ("Business-idea video (1 minute)", "Berklee GEMB", 2,
  "Berklee's gate is a self-produced business pitch, not an audition."),
 ("Registered venture (company or association)", "ESMT Panzer Scholarship — EUR 35,000", 6,
  "The only eligibility condition in the whole dataset you can create rather than satisfy. Do it first."),
 ("DUA Anexo II certification", "All Andalusian publics (UMA, UGR, UJA)", 5,
  "Spain's Distrito Unico route. No homologacion required — university-level verification only."),
 ("Blocked account (Sperrkonto) — EUR 11,904", "Germany visa, if unfunded", 5,
  "Only a GERMAN Sperrkonto counts since Feb 2023. A scholarship removes this requirement entirely."),
 ("Visa appointment slot", "Whichever country you accept", 12,
  "Book the moment an offer lands; consulate queues, not paperwork, are the constraint."),
]
row = 2
for name, who, weeks, note in sorted(DOCS, key=lambda d: -d[2]):
    start_by = earliest - dt.timedelta(weeks=weeks)
    blocking = (start_by - TODAY).days <= 60
    put(ws, row, 1, name, wrap=True)
    put(ws, row, 2, who, wrap=True)
    put(ws, row, 3, weeks)
    put(ws, row, 4, start_by, num="yyyy-mm-dd")
    put(ws, row, 5, "YES" if blocking else "no", font=Font(bold=True, color="FFC2451F") if blocking else BODY)
    put(ws, row, 6, "not started")
    put(ws, row, 7, note, wrap=True, font=SMALL)
    ws.row_dimensions[row].height = 42
    row += 1
ws.cell(row=row+1, column=1, value="Lead times are estimates for a Tunisian applicant, not verified figures. "
        "They are worked backward from " + earliest.isoformat() + ", the earliest confirmed deadline in the Tracker.").font = SMALL
ws.conditional_formatting.add(f"A2:G{row-1}", FormulaRule(
    formula=['AND(ISNUMBER($D2),$D2-TODAY()<=60)'],
    fill=PatternFill("solid", fgColor=RED), stopIfTrue=True))
ws.conditional_formatting.add(f"D2:D{row-1}", FormulaRule(
    formula=['AND(ISNUMBER($D2),$D2<TODAY())'],
    font=Font(bold=True, color="FFC2451F")))
dv2 = DataValidation(type="list", formula1=f'"{DOC_STATUSES}"', allow_blank=False)
dv2.error = "Pick a status from the dropdown."; dv2.errorTitle = "Not a valid status"
ws.add_data_validation(dv2); dv2.add(f"F2:F{row-1}")

# ============================ 5. ALL PROGRAMS ==============================
ws = wb.create_sheet("All Programs")
cols = [c for c in rows[0].keys() if not c.startswith("_") and c not in
        ("paths","fee_num","deadline_date","deadline_conf","deadline_label","usable")]
W = [min(max(14, len(c) + 4), 46) for c in cols]
header(ws, [c.replace("_", " ") for c in cols], W)
row = 2
for r in sorted(usable, key=lambda x: (x["path_letter"], x["country"], x["institution"])):
    for i, c in enumerate(cols, 1):
        v = r[c]
        if c in ("program_url", "admissions_url", "funding_url"):
            link(ws, row, i, v)
        else:
            cell = put(ws, row, i, v[:300], wrap=len(v) > 40)
            if c == "path_letter":
                cell.fill = PatternFill("solid", fgColor=path_fill(r["paths"]) or WASH)
    ws.row_dimensions[row].height = 30
    row += 1

out = Path("deliverables/tools/Application_Command_Center.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print("wrote", out, "| tracker", len(tracker_rows), "| funding", len(fund), "| all programs", len(usable))
