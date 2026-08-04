#!/usr/bin/env python3
"""Build the private-institution workbook for Spain / Portugal / Netherlands / Italy.

The single most important column here is ACCREDITATION. In all four countries a
private school may legally sell a "Master" that is not a recognised degree:
  Spain        Máster Universitario Oficial (ANECA)   vs  Título Propio
  Italy        Laurea magistrale / Diploma Acc. II    vs  "Master" di I livello (60 CFU)
  Portugal     Mestrado (A3ES)                        vs  Pós-graduação
  Netherlands  NVAO-accredited MA/MSc                 vs  unaccredited private course
Only the left-hand column is a real master's degree, and in most cases only the
left-hand column supports a student visa for a non-EU applicant.
"""
import json, re, sys
sys.path.insert(0, '.masters-search')
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = ".masters-search/results/private-es-pt-nl-it.json"
OUT = "results/private-schools-es-pt-nl-it.xlsx"

HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(color="FFFFFF", bold=True, size=11)
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GREY = PatternFill("solid", fgColor="EDEDED")


COUNTRIES = ("Spain", "Portugal", "Netherlands", "Italy")


def country_of(r):
    """Records sometimes carry a sentence in the country field (e.g. a joint
    programme naming all three partners). Take the first country named."""
    raw = (r.get("country") or "").strip()
    for c in COUNTRIES:
        if raw.startswith(c):
            return c
    for c in COUNTRIES:
        if c.lower() in raw.lower():
            return c
    return raw or "Unknown"


def txt(s, n=None):
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    if not s or s.lower() in ("n/a", "na", "none", "unknown", "-"):
        return ""
    return s if (n is None or len(s) <= n) else s[: n - 1] + "…"


def accred_bucket(r):
    """Green / Amber / Red + one-line reason, from the free-text accreditation field."""
    a = (r.get("accreditationStatus") or "").strip()
    low = a.lower()
    if not a or low.startswith("n/a") or low.startswith("not applicable"):
        return "n/a", "Not a relevant programme for this search"
    if low.startswith("official") or re.match(r"^\s*(yes|accredited)", low):
        return "OFFICIAL", "Recognised degree in its own country"
    if low.startswith("not official") or "titulo propio" in low or "título propio" in low:
        return "NOT OFFICIAL", "Private diploma, not a recognised degree"
    if "neche" in low or "united states" in low or "us degree" in low:
        return "FOREIGN", "Real degree, but foreign — needs homologation locally"
    if "mixed" in low or "unverified" in low:
        return "UNCLEAR", "Could not confirm — ask the school in writing"
    return "UNCLEAR", "Could not confirm — ask the school in writing"


def verdict(r):
    v = (r.get("valueVerdict") or "").strip().upper()
    for k in ("WORTH IT", "CONDITIONAL", "AVOID"):
        if v.startswith(k):
            return k
    return ""


def visa(r):
    s = (r.get("visaEligible") or "").strip().lower()
    if s.startswith("yes"):
        return "Yes"
    if s.startswith("no") and not s.startswith("n/a"):
        return "No"
    return "Check"


COLS = [
    ("Country", "country", 14),
    ("City", "city", 22),
    ("Institution", "institution", 40),
    ("Programme", "program", 42),
    ("Accreditation", None, 15),
    ("What that means", None, 44),
    ("Visa OK", None, 9),
    ("Verdict", None, 13),
    ("Qualification awarded", "qualification", 40),
    ("Language", "language", 14),
    ("Length / ECTS", "durationEcts", 18),
    ("Tuition (total)", "tuitionTotal", 26),
    ("Tuition (per year)", "tuitionPerYear", 24),
    ("Other fees", "otherFees", 34),
    ("Equipment you must buy", "equipmentRequired", 30),
    ("Payment plans", "paymentPlans", 28),
    ("ALL-IN estimate", "totalCostEstimate", 34),
    ("Scholarships", "scholarships", 34),
    ("Deadline", "deadline", 26),
    ("Applications open", "applicationOpens", 22),
    ("Intakes", "intakes", 18),
    ("Entry requirements", "entryRequirements", 40),
    ("Accepts non-music BSc", "acceptsNonMusic", 16),
    ("Portfolio required", "portfolioSpec", 36),
    ("Audition / interview", "auditionSpec", 30),
    ("English test", "englishTest", 26),
    ("Facilities", "facilities", 34),
    ("Reputation signals", "reputationSignals", 36),
    ("Recommendation", "recommendation", 50),
    ("Source", "sourceUrl", 30),
]

ORDER = {"OFFICIAL": 0, "FOREIGN": 1, "UNCLEAR": 2, "NOT OFFICIAL": 3, "n/a": 4}
VORDER = {"WORTH IT": 0, "CONDITIONAL": 1, "": 2, "AVOID": 3}


def build_rows(data):
    rows = []
    for r in data:
        b, why = accred_bucket(r)
        rows.append({"_rec": r, "_bucket": b, "_why": why, "_verdict": verdict(r),
                     "_visa": visa(r)})
    rows.sort(key=lambda x: (country_of(x["_rec"]),
                             ORDER.get(x["_bucket"], 5),
                             VORDER.get(x["_verdict"], 2),
                             x["_rec"].get("institution") or ""))
    return rows


def value(row, key, label):
    r = row["_rec"]
    if label == "Country":
        return country_of(r)
    if label == "Accreditation":
        return row["_bucket"]
    if label == "What that means":
        return row["_why"]
    if label == "Visa OK":
        return row["_visa"]
    if label == "Verdict":
        return row["_verdict"] or "—"
    if label == "Accepts non-music BSc":
        s = (r.get("acceptsNonMusic") or "").lower()
        return "Yes" if s.startswith("yes") else ("No" if s.startswith("no") else "Check")
    return txt(r.get(key), 400)


def sheet(wb, title, rows, note=None, wide_cols=None):
    ws = wb.create_sheet(title[:31])
    ridx = 1
    if note:
        ws.cell(1, 1, note).font = Font(bold=True, size=11, color="1F3864")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        ridx = 3
    for c, (label, _, w) in enumerate(COLS, 1):
        cell = ws.cell(ridx, c, label)
        cell.fill, cell.font = HDR, HDRF
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[ridx].height = 30
    for i, row in enumerate(rows):
        rr = ridx + 1 + i
        for c, (label, key, _) in enumerate(COLS, 1):
            cell = ws.cell(rr, c, value(row, key, label))
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        acc = ws.cell(rr, 5)
        acc.fill = {"OFFICIAL": GREEN, "FOREIGN": AMBER, "UNCLEAR": AMBER,
                    "NOT OFFICIAL": RED}.get(row["_bucket"], GREY)
        acc.font = Font(bold=True)
        v = ws.cell(rr, 8)
        v.fill = {"WORTH IT": GREEN, "CONDITIONAL": AMBER, "AVOID": RED}.get(row["_verdict"], GREY)
        src = ws.cell(rr, len(COLS))
        u = row["_rec"].get("sourceUrl") or ""
        if u.startswith("http"):
            src.hyperlink, src.value, src.font = u, "official page", Font(color="0563C1", underline="single")
    ws.auto_filter.ref = f"A{ridx}:{get_column_letter(len(COLS))}{ridx + len(rows)}"
    ws.freeze_panes = ws.cell(ridx + 1, 5)
    return ws


def main():
    data = json.load(open(SRC, encoding="utf-8"))
    rows = build_rows(data)
    wb = Workbook()
    wb.remove(wb.active)

    # README tab
    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 110
    lines = [
        ("PRIVATE INSTITUTIONS — Spain, Portugal, Netherlands, Italy", True),
        ("", False),
        ("Read the ACCREDITATION column before anything else.", True),
        ("OFFICIAL      = a real, recognised master's degree in that country. Safe.", False),
        ("FOREIGN       = a real degree, but from another country's system (e.g. a US degree", False),
        ("                taught in Spain). Valid and often prestigious, but to use it for a", False),
        ("                public job or a PhD locally you must go through homologation.", False),
        ("NOT OFFICIAL  = a 'título propio' / private diploma. It is NOT a master's degree in", False),
        ("                the legal sense, usually will NOT support a student visa on its own,", False),
        ("                and will not be accepted for a PhD. Some are still good training —", False),
        ("                but never pay master's-level money for one without knowing this.", False),
        ("UNCLEAR       = the school does not publish it. Ask them in writing before paying.", False),
        ("", False),
        ("The VERDICT column is the verification agent's judgement of value for money", False),
        ("for YOUR profile (Tunisian, BSc Software Engineering, no music degree, needs a visa).", False),
        ("", False),
        ("Every money column is quoted as the school publishes it. ALL-IN estimate adds", False),
        ("tuition + fees + equipment + living to a single realistic first-year number.", False),
        ("", False),
        ("Tabs: Official Only · Worth It · AVOID · By Country · Everything", False),
    ]
    for i, (t, bold) in enumerate(lines, 1):
        c = ws.cell(i, 1, t)
        if bold:
            c.font = Font(bold=True, size=12, color="1F3864")

    official = [r for r in rows if r["_bucket"] in ("OFFICIAL", "FOREIGN")]
    worth = [r for r in rows if r["_verdict"] == "WORTH IT"]
    avoid = [r for r in rows if r["_verdict"] == "AVOID"]

    sheet(wb, "Official Only", official,
          "Recognised degrees only — these are the ones that support a visa and a future PhD.")
    if worth:
        sheet(wb, "Worth It", worth, "Verified as good value for your profile.")
    if avoid:
        sheet(wb, "AVOID", avoid, "Master's-level price, not a master's-level qualification.")
    sheet(wb, "Everything", rows, "Every record found. Filter with the arrows on the header row.")

    by = defaultdict(list)
    for r in rows:
        by[country_of(r["_rec"])].append(r)
    for country in sorted(by):
        sheet(wb, country, by[country], f"{country} — {len(by[country])} private options found.")

    wb.save(OUT)
    print(f"wrote {OUT}: {len(rows)} rows, {len(official)} official/foreign, "
          f"{len(worth)} worth-it, {len(avoid)} avoid, {len(by)} countries")


if __name__ == "__main__":
    main()
