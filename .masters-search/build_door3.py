#!/usr/bin/env python3
"""Door 3 only, Europe only — every music-production and studio-craft master's in one file.

Combines everything this project has collected in this door across every earlier sweep
with the new Door-3 campaign, so there is one place to look rather than six.

Door 3 means making records: production, studio practice, mixing, mastering, Tonmeister,
sound direction. It deliberately excludes electroacoustic composition, sound art and
sonology — those are Door 1 and live in ARTIST-ROUTE.xlsx.
"""
import sys
from collections import defaultdict, Counter
sys.path.insert(0, ".masters-search")
import build_master as M
import build_artist_workbook as A
import write_master_workbook as W
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = "results/DOOR3.xlsx"
OUT_MD = "results/DOOR3.md"
CHEAP = {"Free", "Under EUR 1.5k/yr", "EUR 1.5k-5k/yr"}

EUROPE = {
    "Germany", "Austria", "Switzerland", "Liechtenstein", "France", "Belgium", "Netherlands",
    "Luxembourg", "Monaco", "Ireland", "United Kingdom", "Italy", "Spain", "Portugal", "Greece",
    "Malta", "Cyprus", "Sweden", "Norway", "Denmark", "Finland", "Iceland", "Estonia", "Latvia",
    "Lithuania", "Poland", "Czech Republic", "Slovakia", "Hungary", "Slovenia", "Croatia",
    "Serbia", "Bosnia and Herzegovina", "Montenegro", "North Macedonia", "Albania", "Romania",
    "Bulgaria", "Turkey", "Moldova", "Ukraine", "Georgia", "Armenia", "Multi-country", "EU-wide",
}

# Cost bands are per year; the region groupings are for the by-region tabs.
REGION = {
    "German-speaking": ["Germany", "Austria", "Switzerland", "Liechtenstein"],
    "France & Benelux": ["France", "Belgium", "Netherlands", "Luxembourg", "Monaco"],
    "Southern Europe": ["Italy", "Spain", "Portugal", "Greece", "Malta", "Cyprus"],
    "UK & Ireland": ["United Kingdom", "Ireland"],
    "Nordics & Baltics": ["Sweden", "Norway", "Denmark", "Finland", "Iceland",
                          "Estonia", "Latvia", "Lithuania"],
    "Central & Eastern": ["Poland", "Czech Republic", "Slovakia", "Hungary", "Slovenia",
                          "Croatia", "Serbia", "Bosnia and Herzegovina", "Montenegro",
                          "North Macedonia", "Albania", "Romania", "Bulgaria", "Turkey",
                          "Moldova", "Ukraine", "Georgia", "Armenia"],
}
OF_REGION = {c: r for r, cs in REGION.items() for c in cs}


def collect():
    verdicts = A.load_verdicts()
    rows = []
    for r in M.main():
        if A.door(r) != "Door 3 — Music production & studio":
            continue
        if r["Country"] not in EUROPE:
            continue
        r["_door"] = "Door 3"
        r["_sub"] = A.subtype(r)
        r["_gate"] = A.gate(r)
        v = verdicts.get((A.norm(r["Institution"]), A.norm(r["Programme / scheme"])))
        r["_verdict"] = (v or {}).get("verdict", "")
        r["_verdictWhy"] = (v or {}).get("verdictWhy", "")
        r["_correction"] = (v or {}).get("correction", "")
        r["_reg"] = OF_REGION.get(r["Country"], "Other")
        rows.append(r)
    return rows


def build_workbook(rows):
    degrees = [r for r in rows if r["Qualification level"] == "Master's degree"]
    notdeg = [r for r in rows if r["Qualification level"] not in ("Master's degree", "Unclear")]
    unclear = [r for r in rows if r["Qualification level"] == "Unclear"]
    noaud = [r for r in degrees if "AUDITION" not in r["_gate"]]
    best = [r for r in noaud if r["Chance for you"] == "Strong" and r["Cost band (you)"] in CHEAP]
    cheap = [r for r in degrees if r["Cost band (you)"] in CHEAP]
    funded = [r for r in degrees if r["Funding"] == "Full"]
    schol = [r for r in degrees if r["Scholarship"] != "—"]
    worth = [r for r in rows if r["_verdict"] == "WORTH IT"]
    cond = [r for r in rows if r["_verdict"] == "CONDITIONAL"]
    avoid = [r for r in rows if r["_verdict"] == "AVOID"]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 112
    lines = [
        ("DOOR 3 — MUSIC PRODUCTION & STUDIO CRAFT, EVERY EUROPEAN OPTION IN ONE FILE", True),
        ("", False),
        (f"{len(rows)} records · {len(degrees)} real master's degrees · {len(notdeg)} not degrees "
         f"· {len(unclear)} unclear.", False),
        ("Everything collected across every sweep this project has run, merged with the new", False),
        ("Door-3 campaign. Nothing was searched twice.", False),
        ("", False),
        ("WHAT IS IN: making records. Production, studio practice, mixing, mastering, Tonmeister,", True),
        ("sound direction, popular-music production, beat-making, DJ and electronic production.", False),
        ("WHAT IS OUT: electroacoustic composition, sound art, installation, sonology.", False),
        ("Those are Door 1 and they live in ARTIST-ROUTE.xlsx.", False),
        ("", False),
        ("COLUMN E - VERIFIED VERDICT. A second agent re-opened the institution's own pages and", True),
        ("judged it FOR YOU. Blank means not yet re-checked - a lead, not a finding.", False),
        ("", False),
        ("COLUMN H - ADMISSION GATE, the column that decides where you can apply.", True),
        ("  GREEN Portfolio / exam / interview -> you can BUILD your way in. Four finished tracks.", False),
        ("  RED   AUDITION -> a live performance audition. You cannot acquire that in time.", False),
        ("  A PRACTICAL PRODUCTION TEST IS NOT AN AUDITION. When a school says 'here is raw", True),
        ("  material, make a production', that is the job, not a barrier. Read column P.", False),
        ("", False),
        ("MONEY FACTS ALREADY SETTLED - apply them when you read the cost column:", True),
        ("  Wallonia + French-speaking Brussels: EUR 5,369/yr flat. Tunisia is NOT exempt.", False),
        ("  Flanders: KASK 8,800 - LUCA 9,824 - KC Brussel 17,501 - KC Antwerpen 25,000.", False),
        ("  France: EUR 3,941/yr differentiated, unless a Ministry of Culture school.", False),
        ("  Norway now charges non-EEA tuition everywhere. Uniarts Helsinki is EUR 28,000/yr.", False),
        ("  Czech-taught study at Czech public universities is free to every nationality.", False),
        ("  CHEVENING IS CLOSED TO YOU for 2027 - it needs two years' post-degree work experience.", True),
        ("  So a UK programme whose only funding route was Chevening is unaffordable in practice.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        c = ws.cell(i, 1, t)
        if b:
            c.font = Font(bold=True, size=12, color="1F3864")

    if worth:
        A.sheet(wb, "✅ VERIFIED WORTH IT", worth,
                "Re-opened on official pages and judged worth an application FOR YOU.")
    if cond:
        A.sheet(wb, "⚠️ VERIFIED CONDITIONAL", cond, "Real, but column F names what to resolve first.")
    A.sheet(wb, "★ BEST BETS", best,
            "Real degree · strong chance · cheap or free · NO audition. Read this tab first.")
    A.sheet(wb, "No audition", noaud,
            "Everything reachable with a PORTFOLIO or a production test rather than a performance audition.")
    A.sheet(wb, "Free or cheap", cheap, "Free, or under EUR 5,000 a year at the rate YOU pay.")
    A.sheet(wb, "Fully funded", funded, "Funding covers tuition and usually a stipend.")
    if schol:
        A.sheet(wb, "Has a scholarship", schol, "A named funding scheme is attached to the programme.")
    A.sheet(wb, "All degrees", degrees, "Every real master's degree in this door, across Europe.")
    if notdeg:
        A.sheet(wb, "NOT a degree — avoid", notdeg,
                "Sold as a master but not one — school diplomas, título propio, Master di I livello.")
    if unclear:
        A.sheet(wb, "Award unclear — ask", unclear,
                "The institution does not say plainly what it awards. Email before applying.")
    if avoid:
        A.sheet(wb, "⛔ VERIFIED AVOID", avoid, "Re-checked and ruled out. Column F says why.")
    A.sheet(wb, "Everything", rows, "All records including rejects. Filter with the header arrows.")

    by = defaultdict(list)
    for r in degrees:
        by[r["_reg"]].append(r)
    for reg in sorted(by, key=lambda k: -len(by[k])):
        A.sheet(wb, reg, by[reg], f"{reg} — {len(by[reg])} degrees.")

    wb.save(OUT)
    return dict(rows=len(rows), degrees=len(degrees), notdeg=len(notdeg), unclear=len(unclear),
                noaud=len(noaud), best=len(best), cheap=len(cheap), funded=len(funded),
                worth=len(worth), cond=len(cond), avoid=len(avoid), by=by)


if __name__ == "__main__":
    rows = collect()
    s = build_workbook(rows)
    print(f"wrote {OUT}")
    print(f"  {s['rows']} records · {s['degrees']} degrees · {s['notdeg']} not degrees · "
          f"{s['unclear']} unclear")
    print(f"  {s['best']} best bets · {s['noaud']} no-audition · {s['cheap']} cheap · "
          f"{s['funded']} funded · verdicts {s['worth']}/{s['cond']}/{s['avoid']}")
    print("  by region :", {k: len(v) for k, v in s["by"].items()})
    print("  by country:", dict(Counter(r["Country"] for r in rows if
                                        r["Qualification level"] == "Master's degree")))
    print("  by subtype:", dict(Counter(r["_sub"] for r in rows if
                                        r["Qualification level"] == "Master's degree")))
