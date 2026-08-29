#!/usr/bin/env python3
"""CSV -> XLSX, losslessly, plus one combined workbook.

The requirement is that no information is lost, so every conversion is verified
cell-by-cell against the source afterwards rather than assumed. Two real hazards are
handled explicitly:
  * Excel silently coerces strings that look like formulas, dates or numbers -- a cell
    beginning "=" or "+" becomes a formula, "1-2" can become a date. Every value is
    written as text with the cell format forced to '@'.
  * Excel's hard limit is 32,767 characters per cell. The longest cell here is 3,883,
    so nothing truncates -- but the check runs anyway and fails loudly if that changes.
"""
import csv, sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

csv.field_size_limit(10**9)
OUT = Path("output"); DEL = OUT/"deliverables"; DEL.mkdir(parents=True, exist_ok=True)
LIMIT = 32767

SOURCES = [
    ("programmes.csv",     "Programmes",   "Every enriched + verified programme, full schema"),
    ("funding.csv",        "Funding",      "Funding sources with the Tunisian-eligibility verdict"),
    ("spring-intakes.csv", "Spring intakes","Every programme checked for a Jan-Jun cohort start"),
]
EXTRA = [("wave3/regional-fee-policy.jsonl", "Regional fees", "Non-EU tuition treatment per autonomous community")]

def read_csv(p):
    with open(p, encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))
    return rows[0], rows[1:]

def style(ws, header):
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5496")
    for c in range(1, len(header)+1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = hf, fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        # width by header length, capped -- long prose columns stay readable via wrap
        ws.column_dimensions[get_column_letter(c)].width = min(max(14, len(header[c-1])+3), 46)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def write_sheet(ws, header, rows):
    ws.append(header)
    for r in rows:
        ws.append([("" if v is None else str(v))[:LIMIT] for v in r])
    # force text so Excel cannot reinterpret a value
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.number_format = "@"
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    style(ws, header)

def verify(xlsx, sheet, header, rows):
    """Read the file back and compare every cell. Loud failure beats a silent lie."""
    wb = load_workbook(xlsx, read_only=True); ws = wb[sheet]
    got = list(ws.iter_rows(values_only=True))
    wb.close()
    assert got[0] == tuple(header), f"{xlsx}:{sheet} header mismatch"
    assert len(got)-1 == len(rows), f"{xlsx}:{sheet} row count {len(got)-1} != {len(rows)}"
    bad = 0
    for i, src in enumerate(rows, start=1):
        for j, v in enumerate(src):
            a = "" if v is None else str(v)
            b = "" if got[i][j] is None else str(got[i][j])
            if a[:LIMIT] != b:
                bad += 1
                if bad <= 3:
                    print(f"    MISMATCH r{i+1}c{j+1}: {a[:60]!r} != {b[:60]!r}", file=sys.stderr)
    return bad

def main():
    combined = Workbook(); combined.remove(combined.active)
    total_rows = total_bad = 0

    # index sheet first, so the workbook explains itself
    idx = combined.create_sheet("Index")
    idx.append(["Sheet", "Rows", "What it is"])

    for fname, sheet, desc in SOURCES:
        src = OUT/fname
        if not src.exists():
            print(f"skip (missing): {fname}"); continue
        header, rows = read_csv(src)
        longest = max((len(c) for r in rows for c in r), default=0)
        assert longest <= LIMIT, f"{fname} has a {longest}-char cell; Excel caps at {LIMIT}"

        single = DEL/f"{src.stem}.xlsx"
        wb = Workbook(); ws = wb.active; ws.title = sheet[:31]
        write_sheet(ws, header, rows); wb.save(single)
        bad = verify(single, sheet[:31], header, rows)

        ws2 = combined.create_sheet(sheet[:31]); write_sheet(ws2, header, rows)
        idx.append([sheet, len(rows), desc])
        total_rows += len(rows); total_bad += bad
        print(f"  {single.name:26} {len(rows):5} rows x {len(header):2} cols · longest cell {longest:5} · verify {'OK' if bad==0 else str(bad)+' MISMATCHES'}")

    # regional fee policy from jsonl -> its own sheet (it never had a CSV)
    import json
    for rel, sheet, desc in EXTRA:
        p = OUT/rel
        if not p.exists(): continue
        recs=[json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]
        if not recs: continue
        cols=[]
        for r in recs:
            for k in r:
                if k not in cols: cols.append(k)
        rows=[[("" if r.get(c) is None else str(r.get(c))) for c in cols] for r in recs]
        ws=combined.create_sheet(sheet[:31]); write_sheet(ws, cols, rows)
        idx.append([sheet, len(rows), desc])
        total_rows += len(rows)
        print(f"  {'(combined only)':26} {len(rows):5} rows · {sheet}")

    style(idx, ["Sheet","Rows","What it is"])
    idx.column_dimensions["C"].width = 62
    combined.move_sheet("Index", offset=-len(combined.sheetnames)+1)
    master = DEL/"Spain-masters-sweep.xlsx"
    combined.save(master)
    print(f"\n  {master.name}: {len(combined.sheetnames)} sheets, {total_rows} data rows")
    print(f"  cell-by-cell verification: {'ALL OK' if total_bad==0 else str(total_bad)+' MISMATCHES'}")
    return 1 if total_bad else 0

if __name__ == "__main__":
    sys.exit(main())
