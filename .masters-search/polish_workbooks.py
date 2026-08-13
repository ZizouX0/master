#!/usr/bin/env python3
"""Make every workbook genuinely filterable and scannable.

Applied to all five workbooks so they behave identically:

  * every data sheet becomes a real Excel TABLE — sticky filter buttons, banded
    rows, and a proper search box inside each filter dropdown
  * freeze panes anchored just below the header and after the identity columns,
    so scrolling right never loses which row you're on
  * duplicate or blank header names repaired (Excel refuses a Table otherwise)
  * uniform row height and no wrapped body text, so a screen shows ~30 rows
    instead of 6

Idempotent: run it as often as you like. Every file is reloaded and checked
afterwards, because a workbook that silently fails to open is worse than one
with a missing filter.
"""
import glob, os, re, sys
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# how many left-hand columns to keep frozen, per workbook
FREEZE_COLS = {
    "MASTER-all-opportunities.xlsx": 3,      # Country · Institution · Programme
    "europe-gapfill.xlsx": 3,                # Country · Region · City
    "private-schools-es-pt-nl-it.xlsx": 4,   # Country · City · Institution · Programme
    "SOUND-and-SCHOLARSHIPS.xlsx": 3,    # Country / Institution / Programme
    "DOOR3.xlsx": 4,                     # Door / Country / Institution / Programme
    "ARTIST-ROUTE.xlsx": 4,              # Door / Country / Institution / Programme
    "sound-design-europe.xlsx": 2,
    "opportunities.xlsx": 2,
}
STYLE = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)


def find_header(ws):
    """The header row is the first row that fills most of the used width and is
    followed by at least one data row."""
    if ws.max_row < 2 or ws.max_column < 3:
        return None
    for r in range(1, min(6, ws.max_row)):
        filled = sum(1 for c in range(1, ws.max_column + 1)
                     if ws.cell(r, c).value not in (None, ""))
        below = sum(1 for c in range(1, ws.max_column + 1)
                    if ws.cell(r + 1, c).value not in (None, ""))
        if filled >= ws.max_column * 0.7 and below >= 2:
            return r
    return None


def fix_headers(ws, hrow):
    """Excel rejects a Table with blank or duplicated header cells."""
    seen, changed = {}, 0
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(hrow, c)
        name = str(cell.value).strip() if cell.value not in (None, "") else ""
        if not name:
            name = f"Column {c}"
            changed += 1
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
            changed += 1
        else:
            seen[name] = 1
        cell.value = name
    return changed


def safe_name(path, sheet, n):
    base = re.sub(r"[^A-Za-z0-9]", "", os.path.basename(path).split(".")[0])[:18]
    tag = re.sub(r"[^A-Za-z0-9]", "", sheet)[:14] or "S"
    return f"T_{base}_{tag}_{n}"


def polish(path):
    wb = load_workbook(path)
    nfrozen = FREEZE_COLS.get(os.path.basename(path), 2)
    touched = 0
    for i, ws in enumerate(wb.worksheets):
        hrow = find_header(ws)
        if hrow is None:
            continue                                   # a README-style sheet
        fix_headers(ws, hrow)

        last_col = get_column_letter(ws.max_column)
        ref = f"A{hrow}:{last_col}{ws.max_row}"

        # a Table supplies its own filter, so drop the sheet-level one
        ws.auto_filter.ref = None
        for existing in list(ws.tables):
            del ws.tables[existing]
        t = Table(displayName=safe_name(path, ws.title, i), ref=ref)
        t.tableStyleInfo = STYLE
        ws.add_table(t)

        ws.freeze_panes = ws.cell(hrow + 1, nfrozen + 1)

        # body rows: single line, top aligned, uniform height
        for r in range(hrow + 1, ws.max_row + 1):
            ws.row_dimensions[r].height = 15
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=False)
        ws.row_dimensions[hrow].height = 32
        touched += 1

    wb.save(path)

    # verify it still opens and the tables survived
    check = load_workbook(path)
    ntab = sum(len(ws.tables) for ws in check.worksheets)
    return touched, ntab, len(check.worksheets)


if __name__ == "__main__":
    files = sys.argv[1:] or sorted(glob.glob("results/*.xlsx"))
    for f in files:
        touched, ntab, nsheets = polish(f)
        print(f"{os.path.basename(f):<38} {nsheets:>2} sheets · {touched} tabled · "
              f"{ntab} tables verified · freeze {FREEZE_COLS.get(os.path.basename(f), 2)} cols")
