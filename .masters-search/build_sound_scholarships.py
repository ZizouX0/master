#!/usr/bin/env python3
"""One workbook with only the sound side of the search, plus all the funding.

The master file carries all 1,091 rows including 335 music-business and 193
adjacent programmes. This strips those out and keeps the three tracks that are
actually about making and engineering sound, then puts every scholarship scheme
alongside them so the money and the programmes sit in one place.
"""
import sys
from collections import defaultdict, Counter
sys.path.insert(0, '.masters-search')
import build_master as M
import write_master_workbook as W
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = "results/SOUND-and-SCHOLARSHIPS.xlsx"

SOUND_TRACKS = ("Sound design / production", "Acoustics / audio engineering",
                "Music & sound technology")
CHEAP = {"Free", "Under EUR 1.5k/yr", "EUR 1.5k-5k/yr"}


def main():
    allrows = M.main()
    sound = [r for r in allrows if r["Track"] in SOUND_TRACKS]
    # Funding schemes are attached to programmes rather than standing alone, so a
    # scholarship tab means "rows carrying a named scheme" — kept across every track,
    # because a scheme found on a business programme may still be open to him.
    schemes = [r for r in allrows if r["Scholarship"] != "—"]

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 108
    lines = [
        ("SOUND & FUNDING — the making-sound half of the search, plus all the money", True),
        ("", False),
        (f"{len(sound)} programmes about sound, and {len(schemes)} rows carrying a named "
         f"scholarship scheme.", False),
        ("", False),
        ("WHAT IS IN, AND WHAT IS NOT", True),
        ("  IN   Sound design / production   — making records: production, mixing, sound design, DJing", False),
        ("  IN   Acoustics / audio engineering — the technical side: signal processing, spatial audio", False),
        ("  IN   Music & sound technology    — game audio, immersive, music tech", False),
        ("  OUT  Music business (335 rows)   — in MASTER-all-opportunities.xlsx if you want it", False),
        ("  OUT  Other / adjacent (193 rows) — media and film degrees where sound is one component", False),
        ("", False),
        ("DO NOT SKIP THE ENGINEERING TRACK. Your two best-rated options - TU Ilmenau and", True),
        ("HfK Bremen - are NOT filed under sound design. They are engineering and digital-media", False),
        ("degrees, which is exactly why they need no portfolio, no audition and no music degree.", False),
        ("", False),
        ("THE FIRST EIGHT COLUMNS ARE THE DECISION. A-C name the programme and stay frozen.", True),
        ("  Chance for you   GREEN Strong - the entry rules name your background, or take any degree", False),
        ("                   AMBER Possible - plausible, but something needs asking or arguing", False),
        ("                   RED   Weak - needs a music degree, audition, experience or a language", False),
        ("  Cost band        tuition per year at the rate YOU pay as a non-EU national.", False),
        ("                   A published whole-programme total is halved to get a per-year figure.", False),
        ("                   'Not published' means the school publishes no figure - not that it is free.", False),
        ("", False),
        ("TABS  Best bets - Sound design & production - Audio engineering - Music tech -", False),
        ("      No portfolio needed - Fully funded - Has a scholarship - Scholarship index - By country", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        c = ws.cell(i, 1, t)
        if b:
            c.font = Font(bold=True, size=12, color="1F3864")

    best = [r for r in sound if r["Chance for you"] == "Strong" and r["Cost band (you)"] in CHEAP]
    nofolio = [r for r in sound
               if not (r["_rec"].get("portfolioSpec") or r["_rec"].get("portfolioReq") or "").strip()
               or (r["_rec"].get("portfolioSpec") or r["_rec"].get("portfolioReq") or "").lower().startswith(("none", "no portfolio", "not required"))]
    funded = [r for r in sound if r["Funding"] == "Full"]
    sfunded = [r for r in schemes if r["Track"] in SOUND_TRACKS]

    W.sheet(wb, "★ BEST BETS", best,
            "Sound programmes that are BOTH a strong chance for you AND cheap or free. Start here.")
    W.sheet(wb, "Sound design & production", [r for r in sound if r["Track"] == "Sound design / production"],
            "Making records — production, mixing, mastering, sound design, electronic music, DJing.")
    W.sheet(wb, "Audio engineering", [r for r in sound if r["Track"] == "Acoustics / audio engineering"],
            "The technical side. This is where YOUR degree is an advantage rather than an exception.")
    W.sheet(wb, "Music tech", [r for r in sound if r["Track"] == "Music & sound technology"],
            "Game audio, immersive and spatial audio, music technology.")
    if nofolio:
        W.sheet(wb, "No portfolio needed", nofolio,
                "No portfolio requirement published — you could apply to these before yours exists.")
    W.sheet(wb, "Fully funded", funded, "Funding covers tuition and usually a stipend.")
    W.sheet(wb, "Sound + a scholarship", sfunded,
            "Sound programmes that carry a named funding scheme.")
    W.sheet(wb, "All scholarships", schemes,
            "Every row carrying a named scheme, across ALL tracks — a scheme found on another "
            "programme may still be open to you.")
    W.sheet(wb, "All sound programmes", sound, "Every sound, audio and music-technology programme found.")

    # one row per scheme, so the funding landscape is visible on its own
    idx = defaultdict(list)
    for r in schemes:
        idx[r["Scholarship"]].append(r)
    ws = wb.create_sheet("Scholarship index")
    for i, (h, w) in enumerate([("Scholarship scheme", 60), ("Programmes", 11),
                                ("Of those, sound", 14), ("Countries", 40),
                                ("Best chance offered", 18)], 1):
        c = ws.cell(1, i, h)
        c.fill, c.font = W.HDR, W.HDRF
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    order = {"Strong": 0, "Possible": 1, "Weak": 2}
    for n, (name, rs) in enumerate(sorted(idx.items(), key=lambda kv: -len(kv[1])), 2):
        bestch = sorted((x["Chance for you"] for x in rs), key=lambda c: order[c])[0]
        ws.cell(n, 1, name)
        ws.cell(n, 2, len(rs))
        ws.cell(n, 3, sum(1 for x in rs if x["Track"] in SOUND_TRACKS))
        ws.cell(n, 4, ", ".join(sorted({x["Country"] for x in rs}))[:250])
        c = ws.cell(n, 5, bestch)
        c.fill = {"Strong": W.GREEN, "Possible": W.AMBER, "Weak": W.RED}[bestch]
    ws.auto_filter.ref = f"A1:E{1 + len(idx)}"
    ws.freeze_panes = ws.cell(2, 1)

    by = defaultdict(list)
    for r in sound:
        by[r["Country"]].append(r)
    for c in sorted(by, key=lambda k: -len(by[k]))[:14]:
        W.sheet(wb, c, by[c], f"{c} — {len(by[c])} sound programmes.")

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(sound)} sound programmes · {len(best)} best bets · {len(nofolio)} no-portfolio · "
          f"{len(funded)} fully funded · {len(schemes)} with a scheme · {len(idx)} distinct schemes")
    print("  by track:", dict(Counter(r["Track"] for r in sound)))


if __name__ == "__main__":
    main()
