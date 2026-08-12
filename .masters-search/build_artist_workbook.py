#!/usr/bin/env python3
"""The artist-route workbook: Door 1 and Door 3 only.

Door 1  Electronic music / computer music / sound art / electroacoustic composition
Door 3  Music production / studio craft

These are the two the candidate chose. Everything else — audio engineering, media art,
game audio, music business — is deliberately absent; it lives in the master file.

The columns are ordered for the decision he actually faces, which is not cost first.
It is: is this a real degree, will they let me in without a music degree, and what
exactly must I submit. So the admission gate sits next to the programme name, and the
portfolio specification is a first-class column rather than something buried at column 29.
"""
import re, sys
from collections import defaultdict, Counter
sys.path.insert(0, ".masters-search")
import build_master as M
import write_master_workbook as W
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT = "results/ARTIST-ROUTE.xlsx"
CHEAP = {"Free", "Under EUR 1.5k/yr", "EUR 1.5k-5k/yr"}

# --- which door a programme belongs to --------------------------------------------
# Door 1 is matched first: a programme called "Electronic Music Production" is an
# electronic-music degree that happens to teach production, not a studio-craft degree.
D1 = re.compile(
    r"electronic music|elektronische musik|musica elettronica|m[uú]sica electr[oó]nica|"
    r"muzyka elektroniczna|elektronick[aá] hudba|elektronikus zene|elektronisk musik|"
    r"elektroninen musiikki|elektronik m[uü]zik|elektronska muzika|"
    r"computer music|computermusik|informatique musicale|informatica musicale|"
    r"sonolog|sound art|klangkunst|arte sonora|art sonore|arts sonores|sonic art|"
    r"electroacoust|elektroakust|[eé]lectroacoust|acousmat|akusmat|"
    r"composition|komposition|composizion|composici[oó]n|composi[cç][aã]o|kompozycja|"
    r"kompozic|zeneszerz|skladba|"
    r"experimental music|musique exp[eé]rimentale|nuove tecnologie|"
    r"creation musicale|cr[eé]ation sonore|new media.*sound|sound.*new media|"
    r"multimedia composition|multimediale komposition|medienkomposition|"
    r"music technology|musiktechnologie|musikteknolog|m[uü]zik teknoloji", re.I)
D3 = re.compile(
    r"music production|musikproduktion|musikkproduksjon|muziekproductie|"
    r"produzione musicale|producci[oó]n musical|produ[cç][aã]o musical|"
    r"produkcja muzyczna|muzi[čc]ka produkcija|zenei produkci|musiikkituotanto|"
    r"record(?:ing)? production|studio production|record production|"
    r"mixing|mastering|beatmaking|\bdj\b|turntabl|"
    r"tonmeister|musikregie|realizacja d[źz]wi[eę]ku|zvukov[aá] tvorba|"
    r"r[eé]alisation sonore|m[eé]tiers du son|master en son|popular music production|"
    r"popularmusik", re.I)

SUBTYPE = [
    ("Sonology / computer music", r"sonolog|computer music|computermusik|informatique musicale|informatica musicale|musikinformatik"),
    ("Electroacoustic / acousmatic", r"electroacoust|elektroakust|[eé]lectroacoust|acousmat|akusmat"),
    ("Sound art / installation", r"sound art|klangkunst|arte sonora|arts? sonore|sonic art|installation"),
    ("Electronic music composition", r"electronic music|elektronische musik|musica elettronica|m[uú]sica electr[oó]nica|muzyka elektroniczna|elektronik m[uü]zik|elektronisk musik|elektronick[aá] hudba"),
    ("Composition with technology", r"composition|komposition|composizion|composici[oó]n|composi[cç][aã]o|kompozycja|kompozic|zeneszerz|skladba"),
    ("Music technology", r"music technology|musiktechnologie|musikteknolog|m[uü]zik teknoloji|nuove tecnologie"),
    ("Studio / record production", r"music production|musikproduktion|produzione musicale|producci[oó]n musical|produ[cç][aã]o musical|produkcja muzyczna|muzi[čc]ka produkcija|record(?:ing)? production|studio production|muziekproductie"),
    ("Mixing / mastering", r"mixing|mastering"),
    ("Tonmeister / sound direction", r"tonmeister|musikregie|realizacja d[źz]wi[eę]ku|zvukov[aá] tvorba|r[eé]alisation sonore|m[eé]tiers du son|master en son"),
    ("DJ / performance", r"\bdj\b|turntabl|live electronics|performance"),
]


def hay(r):
    return (f"{r['Programme / scheme']} "
            f"{M.txt(r['_rec'].get('whatYouStudy') or r['_rec'].get('focus'), 300)}")


def door(r):
    t = hay(r)
    if D1.search(t):
        return "Door 1 — Electronic music & sound art"
    if D3.search(t):
        return "Door 3 — Music production & studio"
    return None


def subtype(r):
    t = hay(r)
    for name, rx in SUBTYPE:
        if re.search(rx, t, re.I):
            return name
    return "Other"


# --- what actually stands between him and a place ----------------------------------
RX_AUD = re.compile(r"live audition|in[- ]person audition|instrumental audition|"
                    r"audition (?:is )?(?:required|mandatory)|esame di ammissione|"
                    r"prueba de acceso|[eé]preuve d'admission|vorspiel", re.I)
RX_NOAUD = re.compile(r"no live (?:admission )?exam|no audition|no live audition|"
                      r"audition: none|selection is (?:entirely )?(?:document|portfolio)|"
                      # "YES, but NOT an instrumental audition" — an entrance test that is a
                      # production or coding task is not the barrier a performance audition is
                      r"not an? (?:instrumental|performance|live|music|classical)[- ]?\w* ?audition|"
                      r"not? (?:instrumental )?audition|"
                      r"does not test (?:classical )?instrumental|practical production test", re.I)
RX_EXAM = re.compile(r"entrance exam|written exam|aptitude test|eignungspr[uü]fung|"
                     r"admission interview|interview", re.I)
RX_PF = re.compile(r"portfolio|works? must be submitted|representative works|"
                   r"submit (?:at least )?\w+ (?:works|pieces|compositions)|"
                   r"compositions|demo|showreel", re.I)


def gate(r):
    p = str(r["_rec"].get("portfolioSpec") or r["_rec"].get("portfolioReq") or "")
    a = str(r["_rec"].get("auditionSpec") or "")
    e = str(r["_rec"].get("entryRequirements") or "")
    both = f"{p} {a} {e}"
    has_pf = bool(RX_PF.search(both)) or p.strip().upper().startswith("Y")
    # a stated "no audition" beats a generic audition word appearing in the same text
    aud = bool(RX_AUD.search(both)) and not RX_NOAUD.search(both)
    exam = bool(RX_EXAM.search(both))
    if aud:
        return "AUDITION — hardest for you"
    if has_pf and exam:
        return "Portfolio + exam/interview"
    if has_pf:
        return "Portfolio only"
    if exam:
        return "Exam/interview only"
    if not both.strip():
        return "Not published — ask them"
    return "None found in the text"


GATE_FILL = {
    "Portfolio only": W.GREEN,
    "Portfolio + exam/interview": W.GREEN,
    "Exam/interview only": W.GREEN,
    "None found in the text": W.GREEN,
    "AUDITION — hardest for you": W.RED,
    "Not published — ask them": W.GREY,
}

COLS = [
    ("Door", lambda r: r["_door"], 30),
    ("Country", lambda r: r["Country"], 16),
    ("Institution", lambda r: r["Institution"], 42),
    ("Programme", lambda r: r["Programme / scheme"], 44),
    ("Verified verdict", lambda r: r["_verdict"] or "—", 16),
    ("Why that verdict", lambda r: M.txt(r["_verdictWhy"], 600), 56),
    ("Correction found", lambda r: M.txt(r["_correction"], 400), 44),
    ("Admission gate", lambda r: r["_gate"], 26),
    ("Qualification level", lambda r: r["Qualification level"], 22),
    ("Chance for you", lambda r: r["Chance for you"], 14),
    ("Cost band (you)", lambda r: r["Cost band (you)"], 16),
    ("Taught in", lambda r: r["Taught in"], 16),
    ("Funding", lambda r: r["Funding"], 10),
    ("Sub-type", lambda r: r["_sub"], 26),
    ("Portfolio — exactly what", lambda r: W.g(r, "portfolioSpec", "portfolioReq", n=600), 56),
    ("Audition / exam", lambda r: W.g(r, "auditionSpec", n=400), 40),
    ("Accepts non-music degree", lambda r: W.g(r, "acceptsNonMusic", n=400), 40),
    ("Entry requirements", lambda r: W.g(r, "entryRequirements", n=600), 50),
    ("Deadline", lambda r: W.g(r, "deadline", n=300), 34),
    ("Applications open", lambda r: W.g(r, "applicationOpens"), 26),
    ("YOUR tuition", lambda r: W.g(r, "tuitionNonEU", "tuitionIntl", "tuitionTotal", "tuitionPerYear"), 36),
    ("All-in estimate", lambda r: W.g(r, "totalCostEstimate"), 32),
    ("Scholarship", lambda r: r["Scholarship"], 30),
    ("Scholarship detail", lambda r: W.g(r, "scholarships", "scholarship", "stipendCoverage"), 40),
    ("Why that chance", lambda r: r["Why that chance"], 44),
    ("What you study", lambda r: W.g(r, "whatYouStudy", "focus"), 44),
    ("Length / ECTS", lambda r: W.g(r, "durationEcts", n=80), 18),
    ("Language certificate", lambda r: W.g(r, "languageRequirement"), 28),
    ("English test", lambda r: W.g(r, "englishTest"), 24),
    ("Public/Private", lambda r: r["Public/Private"], 13),
    ("City", lambda r: r["City"], 18),
    ("Accreditation", lambda r: W.g(r, "accreditationStatus"), 30),
    ("Verdict", lambda r: W.g(r, "valueVerdict", n=20) or "—", 12),
    ("Recommendation", lambda r: W.g(r, "recommendation", "fitNotes", n=700), 56),
    ("Found by", lambda r: r["_rec"].get("_source", ""), 22),
    ("Source", lambda r: r["_rec"].get("sourceUrl") or "", 26),
]
IDX = {n: i for i, (n, _f, _w) in enumerate(COLS, 1)}


def sheet(wb, title, rows, note=None):
    ws = wb.create_sheet(title[:31])
    top = 1
    if note:
        c = ws.cell(1, 1, note)
        c.font = Font(bold=True, size=11, color="1F3864")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        top = 3
    for i, (label, _, w) in enumerate(COLS, 1):
        c = ws.cell(top, i, label)
        c.fill, c.font = W.HDR, W.HDRF
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[top].height = 30
    for n, r in enumerate(rows):
        rr = top + 1 + n
        for i, (_, fn, _w) in enumerate(COLS, 1):
            ws.cell(rr, i, fn(r)).alignment = Alignment(vertical="top", wrap_text=False)
        ws.cell(rr, IDX["Admission gate"]).fill = GATE_FILL.get(r["_gate"], W.GREY)
        vc = ws.cell(rr, IDX["Verified verdict"])
        vc.fill = {"WORTH IT": W.GREEN, "CONDITIONAL": W.AMBER, "AVOID": W.RED}.get(
            r["_verdict"], W.GREY)
        if r["_verdict"]:
            vc.font = Font(bold=True)
        ws.cell(rr, IDX["Qualification level"]).fill = (
            W.GREEN if r["Qualification level"] == "Master's degree" else W.RED)
        ws.cell(rr, IDX["Chance for you"]).fill = {
            "Strong": W.GREEN, "Possible": W.AMBER, "Weak": W.RED}[r["Chance for you"]]
        ws.cell(rr, IDX["Chance for you"]).font = Font(bold=True)
        ws.cell(rr, IDX["Cost band (you)"]).fill = {
            "Free": W.GREEN, "Under EUR 1.5k/yr": W.GREEN, "EUR 1.5k-5k/yr": W.GREEN,
            "EUR 5k-15k/yr": W.AMBER, "Over EUR 15k/yr": W.RED}.get(r["Cost band (you)"], W.GREY)
        ws.cell(rr, IDX["Taught in"]).fill = W.GREEN if M.language_ok(r["Taught in"]) else W.GREY
        ws.cell(rr, IDX["Funding"]).fill = {"Full": W.GREEN, "Partial": W.AMBER}.get(r["Funding"], W.GREY)
        src = ws.cell(rr, len(COLS))
        u = r["_rec"].get("sourceUrl") or ""
        if u.startswith("http"):
            src.hyperlink, src.value = u, "official page"
            src.font = Font(color="0563C1", underline="single")
    ws.auto_filter.ref = f"A{top}:{get_column_letter(len(COLS))}{top + len(rows)}"
    ws.freeze_panes = ws.cell(top + 1, 5)     # Door / Country / Institution / Programme
    return ws


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())[:26]


def load_verdicts():
    """The re-verification pass, keyed loosely on institution + programme.

    The verifiers were given the OLD dataset, so their targets span several source
    files and the strings do not match exactly. Institution prefix plus programme
    prefix is enough: within one institution, two programmes in the same door rarely
    share a 26-character prefix.
    """
    import glob as _g, json as _j
    out = {}
    for f in _g.glob(".masters-search/results/artist-sweep/verify-*.json"):
        if "verify-targets" in f:
            continue
        for v in _j.load(open(f, encoding="utf-8")):
            k = (norm(v.get("institution")), norm(v.get("programme")))
            if v.get("verdict"):
                out[k] = v
    return out


def main():
    verdicts = load_verdicts()
    rows = []
    for r in M.main():
        d = door(r)
        if not d:
            continue
        r["_door"] = d
        r["_sub"] = subtype(r)
        r["_gate"] = gate(r)
        v = verdicts.get((norm(r["Institution"]), norm(r["Programme / scheme"])))
        r["_verdict"] = (v or {}).get("verdict", "")
        r["_verdictWhy"] = (v or {}).get("verdictWhy", "")
        r["_correction"] = (v or {}).get("correction", "")
        rows.append(r)
    nv = sum(1 for r in rows if r["_verdict"])
    print(f"  verdicts attached: {nv} of {len(verdicts)} available")

    degrees = [r for r in rows if r["Qualification level"] == "Master's degree"]
    d1 = [r for r in degrees if r["_door"].startswith("Door 1")]
    d3 = [r for r in degrees if r["_door"].startswith("Door 3")]
    notdeg = [r for r in rows if r["Qualification level"] != "Master's degree"]
    noaud = [r for r in degrees if "AUDITION" not in r["_gate"]]
    best = [r for r in noaud if r["Chance for you"] == "Strong" and r["Cost band (you)"] in CHEAP]
    cheap = [r for r in degrees if r["Cost band (you)"] in CHEAP]
    funded = [r for r in degrees if r["Funding"] == "Full"]
    schol = [r for r in degrees if r["Scholarship"] != "—"]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 112
    lines = [
        ("THE ARTIST ROUTE — the only two doors you chose, and nothing else", True),
        ("", False),
        (f"{len(degrees)} real master's degrees. Door 1: {len(d1)} · Door 3: {len(d3)}.", False),
        (f"{len(notdeg)} more records that are NOT degrees are kept on their own tab so you can check them.", False),
        ("", False),
        ("DOOR 1 - ELECTRONIC MUSIC & SOUND ART", True),
        ("  Sonology, computer music, electroacoustic and acousmatic composition, sound art,", False),
        ("  electronic music composition, music technology. You compose and perform your own work,", False),
        ("  with a supervisor and a studio, and you leave with pieces that were played in public.", False),
        ("  THIS IS THE DOOR THAT MAKES YOU AN ARTIST.", True),
        ("", False),
        ("DOOR 3 - MUSIC PRODUCTION & STUDIO", True),
        ("  Record production, studio craft, mixing, mastering, Tonmeister, sound direction.", False),
        ("  The actual record-making skills. Narrower and more competitive, and fewer of them", False),
        ("  are affordable - but it is the most directly useful to a producer.", False),
        ("", False),
        ("COLUMNS E-G - THE VERIFICATION PASS. A second agent re-opened the institution's own", True),
        ("pages for 103 of these and judged them FOR YOU: WORTH IT / CONDITIONAL / AVOID, with", False),
        ("the reason in column F and any correction it found in column G. Where column E is blank", False),
        ("the record has not been re-verified - treat it as a lead, not a finding.", False),
        ("", False),
        ("TWO CORRECTIONS THE VERIFIERS FOUND THAT THE COST COLUMN STILL SHOWS WRONG:", True),
        ("  Kunstuni Linz is NOT free. Austria charges third-country nationals EUR 726.72 per", False),
        ("  semester - about EUR 1,505/yr including the OeH fee. Still affordable, but not zero.", False),
        ("  Sonology (The Hague) and DIGICREA each appear THREE TIMES under slightly different", False),
        ("  institution names. Each is ONE programme and ONE application, not three.", False),
        ("", False),
        ("COLUMN H - ADMISSION GATE. This is the column that decides where you can apply.", True),
        ("  GREEN Portfolio only / Portfolio + exam / Exam only  -> you can BUILD your way in.", False),
        ("        Four finished pieces by December 2026 and these open.", False),
        ("  RED   AUDITION  -> needs conservatoire performance training. You cannot acquire this", False),
        ("        in time. Filter these out unless something else makes them worth it.", False),
        ("  GREY  Not published -> email and ask. Do not assume it means no portfolio.", False),
        ("", False),
        ("A portfolio is NOT an audition. Almost none of these programmes need a music degree", True),
        ("either - column N quotes each page on exactly that question. Read it before ruling", False),
        ("anything out on the grounds that you did software engineering.", False),
        ("", False),
        ("The PORTFOLIO column gives the EXACT specification where the institution published one -", False),
        ("how many works, how long, what format. That is your build brief. Sort by it.", False),
        ("", False),
        ("TABS  Best bets - Door 1 - Door 3 - No audition - Free or cheap - Fully funded -", False),
        ("      Has a scholarship - Scholarship index - NOT a degree - Everything - by country", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        c = ws.cell(i, 1, t)
        if b:
            c.font = Font(bold=True, size=12, color="1F3864")

    worth = [r for r in rows if r["_verdict"] == "WORTH IT"]
    cond = [r for r in rows if r["_verdict"] == "CONDITIONAL"]
    avoid = [r for r in rows if r["_verdict"] == "AVOID"]

    if worth:
        sheet(wb, "✅ VERIFIED WORTH IT", worth,
              "Re-opened on the institution's own pages by a second agent and judged worth an "
              "application FOR YOU. The strongest evidence in the whole file.")
    if cond:
        sheet(wb, "⚠️ VERIFIED CONDITIONAL", cond,
              "Real, but something must be resolved first — column F names it.")
    if avoid:
        sheet(wb, "⛔ VERIFIED AVOID", avoid,
              "Re-checked and ruled out for you. Column F says why. Do not spend an application here.")
    sheet(wb, "★ BEST BETS", best,
          "Real degree · strong chance · cheap or free · NO audition. If you read one tab, read this one.")
    sheet(wb, "DOOR 1 — electronic & sound art", d1,
          "Sonology, computer music, electroacoustic composition, sound art, music technology. "
          "The door that develops you as an artist.")
    sheet(wb, "DOOR 3 — production & studio", d3,
          "Record production, studio craft, mixing, mastering, Tonmeister, sound direction.")
    sheet(wb, "No audition", noaud,
          "Everything you can reach with a PORTFOLIO rather than a conservatoire audition.")
    sheet(wb, "Free or cheap", cheap, "Free, or under EUR 5,000 a year at the rate YOU pay.")
    sheet(wb, "Fully funded", funded, "Funding covers tuition and usually a stipend.")
    sheet(wb, "Has a scholarship", schol, "A named funding scheme is attached to the programme.")
    if notdeg:
        sheet(wb, "NOT a degree — avoid", notdeg,
              "In these two fields but NOT a master's degree — título propio, Master di I livello, "
              "certificates. Kept so you can check them, not apply to them.")
    sheet(wb, "Everything", rows, "Both doors, every record. Filter with the arrows on the header row.")

    idx = defaultdict(list)
    for r in schol:
        idx[r["Scholarship"]].append(r)
    ws = wb.create_sheet("Scholarship index")
    for i, (h, w) in enumerate([("Scholarship scheme", 62), ("Programmes", 12),
                                ("Countries", 40), ("Best chance offered", 18)], 1):
        c = ws.cell(1, i, h)
        c.fill, c.font = W.HDR, W.HDRF
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    order = {"Strong": 0, "Possible": 1, "Weak": 2}
    for n, (name, rs) in enumerate(sorted(idx.items(), key=lambda kv: -len(kv[1])), 2):
        bc = sorted((x["Chance for you"] for x in rs), key=lambda c: order[c])[0]
        ws.cell(n, 1, name)
        ws.cell(n, 2, len(rs))
        ws.cell(n, 3, ", ".join(sorted({x["Country"] for x in rs}))[:250])
        ws.cell(n, 4, bc).fill = {"Strong": W.GREEN, "Possible": W.AMBER, "Weak": W.RED}[bc]
    ws.auto_filter.ref = f"A1:D{1 + len(idx)}"
    ws.freeze_panes = ws.cell(2, 1)

    by = defaultdict(list)
    for r in degrees:
        by[r["Country"]].append(r)
    for c in sorted(by, key=lambda k: -len(by[k]))[:14]:
        sheet(wb, c, by[c], f"{c} — {len(by[c])} degrees across both doors.")

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(rows)} records · {len(degrees)} real degrees "
          f"(Door 1: {len(d1)} · Door 3: {len(d3)}) · {len(notdeg)} not degrees")
    print(f"  {len(best)} best bets · {len(noaud)} no-audition · {len(cheap)} cheap · "
          f"{len(funded)} fully funded · {len(idx)} schemes")
    print("  gates :", dict(Counter(r["_gate"] for r in degrees)))
    print("  subtype:", dict(Counter(r["_sub"] for r in degrees)))


if __name__ == "__main__":
    main()
