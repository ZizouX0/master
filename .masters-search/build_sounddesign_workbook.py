#!/usr/bin/env python3
"""Build a browsable Excel workbook for the Europe sound-design/production sweep."""
import json, re, sys
sys.path.insert(0, '.masters-search')
import cost_model
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

SRC = ".masters-search/results/sounddesign-europe.json"
OUT = "results/sound-design-europe.xlsx"

COUNTRY_FIX = {"usa":"USA","uk":"United Kingdom","türkiye":"Turkey","turkiye":"Turkey",
               "czechia":"Czech Republic","the netherlands":"Netherlands"}

def clean_country(raw):
    s = (raw or "").strip()
    if not s: return "Unknown"
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.split(r"[/,;]| \+ | and ", s)[0]
    s = re.sub(r"\s+", " ", s).strip(" -+")
    return COUNTRY_FIX.get(s.lower(), s) or "Unknown"

def short(s, n=90):
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    return s if len(s) <= n else s[:n-1] + "…"

def rating(p):
    r = (p.get("rating") or "").strip().upper()
    return r[0] if r[:1] in ("A","B","C") else "—"

def funding(p):
    s = (p.get("fundingLevel") or "").strip().lower()
    if s.startswith("full"): return "Full"
    if s.startswith("partial"): return "Partial"
    if s.startswith("none"): return "None"
    return "Check"

def yesno(v):
    s = (v or "").strip().lower()
    if s.startswith("yes"): return "Yes"
    if s.startswith("no") and not s.startswith("n/a"): return "No"
    return "Check"

def audition(p):
    t = ((p.get("auditionSpec") or "") + " " + (p.get("entryRequirements") or "")).lower()
    if re.search(r"^none|\bno live|no audition|no on-site", t.strip()): return "No"
    if re.search(r"on-site|in person|in-person|live (audition|performance|test)|aptitude test", t): return "LIVE"
    if re.search(r"online interview|interview|video", t): return "Online"
    if re.search(r"audition|entrance exam", t): return "Check"
    return "None"

def language(p):
    s = (p.get("language") or "").lower()
    if "english" in s: return "English"
    if "french" in s or "français" in s: return "French"
    for L in ("german","italian","spanish","portuguese","polish","czech","dutch","danish",
              "swedish","norwegian","finnish","romanian","turkish","greek","hungarian","slovak"):
        if L in s: return L.capitalize()
    return "Check"

data = json.load(open(SRC, encoding="utf-8"))
rows = []
for p in data:
    c = clean_country(p.get("country"))
    fnd = funding(p)
    # reuse the cost model (field names differ slightly in this dataset)
    est = cost_model.estimate({"Country": c, "Funding": fnd},
                              {"tuitionIntl": p.get("tuitionIntl"),
                               "stipendCoverage": p.get("stipend") or p.get("scholarship")})
    rows.append({
        "Rating": rating(p),
        "Programme": short(p.get("program"), 80),
        "Institution": short(p.get("university"), 60),
        "City": short(p.get("city"), 34),
        "Country": c,
        "Focus": short(p.get("focus"), 40),
        "Language": language(p),
        "Funding": fnd,
        "Est. YOUR cost/yr (EUR)": est["Est. YOUR cost/yr (EUR)"],
        "Accepts non-music?": yesno(p.get("acceptsNonMusic")),
        "Audition": audition(p),
        "Portfolio (summary)": short(p.get("portfolioSpec"), 120),
        "Deadline": short(p.get("deadline"), 70),
        "Verified": (p.get("verified") or "—"),
        "Link": p.get("sourceUrl", ""),
        # detail-only
        "Degree": short(p.get("degree"), 40),
        "Type": short(p.get("institutionType"), 40),
        "Duration/ECTS": short(p.get("durationEcts"), 60),
        "Tuition (published)": short(p.get("tuitionIntl"), 150),
        "Scholarship": short(p.get("scholarship"), 150),
        "Stipend": short(p.get("stipend"), 150),
        "Portfolio (full spec)": short(p.get("portfolioSpec"), 400),
        "Audition (full)": short(p.get("auditionSpec"), 250),
        "Accepts non-music (wording)": short(p.get("acceptsNonMusic"), 300),
        "Entry requirements": short(p.get("entryRequirements"), 400),
        "What you study": short(p.get("whatYouStudy"), 500),
        "Facilities": short(p.get("facilities"), 300),
        "Opens": short(p.get("applicationOpens"), 60),
        "Fit notes": short(p.get("fitNotes"), 300),
        "Verification notes": short(p.get("verificationNotes"), 400),
    })

ORDER = {"A":0, "B":1, "C":2, "—":3}
def costkey(r):
    v = r["Est. YOUR cost/yr (EUR)"]
    return v if isinstance(v, int) else 10**7
rows.sort(key=lambda r: (ORDER[r["Rating"]], {"Full":0,"Partial":1,"Check":2,"None":3}[r["Funding"]], costkey(r)))

MAIN = ["Rating","Programme","Institution","City","Country","Focus","Language","Funding",
        "Est. YOUR cost/yr (EUR)","Accepts non-music?","Audition","Portfolio (summary)",
        "Deadline","Verified","Link"]
DETAIL = MAIN[:-1] + ["Degree","Type","Duration/ECTS","Tuition (published)","Scholarship","Stipend",
        "Portfolio (full spec)","Audition (full)","Accepts non-music (wording)","Entry requirements",
        "What you study","Facilities","Opens","Fit notes","Verification notes","Link"]
W = {"Rating":8,"Programme":56,"Institution":40,"City":22,"Country":16,"Focus":26,"Language":11,
     "Funding":11,"Est. YOUR cost/yr (EUR)":15,"Accepts non-music?":14,"Audition":10,
     "Portfolio (summary)":58,"Deadline":40,"Verified":11,"Link":34,"Degree":24,"Type":26,
     "Duration/ECTS":30,"Tuition (published)":58,"Scholarship":58,"Stipend":58,
     "Portfolio (full spec)":80,"Audition (full)":60,"Accepts non-music (wording)":70,
     "Entry requirements":80,"What you study":90,"Facilities":60,"Opens":26,"Fit notes":70,
     "Verification notes":80}
CENTER = {"Rating","Language","Funding","Est. YOUR cost/yr (EUR)","Accepts non-music?","Audition","Verified","Country"}

NAVY="10233F"; BLUE="2F6DF6"
hf = Font(bold=True, color="FFFFFF", size=10); hfill = PatternFill("solid", fgColor=NAVY)
GREEN=PatternFill("solid",fgColor="C6EFCE"); YELLOW=PatternFill("solid",fgColor="FFF2CC")
GREY=PatternFill("solid",fgColor="F2F2F2"); RED=PatternFill("solid",fgColor="FFC7CE")

def sheet(wb, title, recs, cols=MAIN, tab=None):
    ws = wb.create_sheet(title)
    if tab: ws.sheet_properties.tabColor = tab
    ws.append(cols)
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i); cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for r in recs: ws.append([r.get(c, "") for c in cols])
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = W.get(c, 20)
        for cell in ws[get_column_letter(i)][1:]:
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if c in CENTER else "left")
    if "Link" in cols:
        ci = cols.index("Link") + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci); u = str(cell.value or "")
            if u.startswith("http"):
                cell.hyperlink = u; cell.value = "open ↗"
                cell.font = Font(color=BLUE, underline="single", size=10)
    for col, rules in (("Rating", [("A",GREEN),("B",YELLOW),("C",RED)]),
                       ("Funding", [("Full",GREEN),("Partial",YELLOW),("None",GREY)]),
                       ("Accepts non-music?", [("Yes",GREEN),("No",RED)]),
                       ("Audition", [("No",GREEN),("None",GREEN),("LIVE",RED)])):
        if col in cols:
            L = get_column_letter(cols.index(col) + 1); rng = f"{L}2:{L}{ws.max_row}"
            for val, fill in rules:
                ws.conditional_formatting.add(rng, CellIsRule(operator="equal",
                                              formula=[f'"{val}"'], fill=fill))
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    return ws

wb = Workbook(); wb.remove(wb.active)

# START HERE
ws = wb.create_sheet("START HERE"); ws.sheet_properties.tabColor = BLUE
ws.cell(row=1, column=1, value="Sound Design & Music Production Masters in Europe").font = Font(bold=True, size=16, color=NAVY)
nA = sum(1 for r in rows if r["Rating"]=="A"); nB = sum(1 for r in rows if r["Rating"]=="B")
nC = sum(1 for r in rows if r["Rating"]=="C")
nfull = sum(1 for r in rows if r["Funding"]=="Full"); nnm = sum(1 for r in rows if r["Accepts non-music?"]=="Yes")
guide = ["", "TABS",
 "  A-LIST              The 13 best: excellent fit AND you can realistically get in. START HERE.",
 "  All Programmes      All 319, sorted A -> B -> C, then by funding, then by cost.",
 "  No Audition         Programmes with no live audition — the ones open to you without performance training.",
 "  Accepts Non-Music   Every programme that takes a software-engineering/non-music bachelor.",
 "  Fully Funded        Full-funding programmes only.",
 "  By Country          How many programmes per country, and how many are A-rated.",
 "  Full Detail         Everything: full portfolio spec, entry requirements, curriculum, facilities.",
 "", "HOW TO USE IT",
 "  Row 1 has filter arrows. Click one, untick 'Select All', tick what you want.",
 "  Example — funded options you can enter:  Rating -> A,  Accepts non-music? -> Yes.",
 "  Example — no performance test:  Audition -> No / None.",
 "  Example — French-language routes:  Language -> French  (your French is an advantage here).",
 "", "COLUMNS",
 "  Rating              A = excellent fit and attainable · B = good with a caveat · C = weak or blocked.",
 "  Accepts non-music?  Does the official page accept a non-music bachelor like yours?",
 "  Audition            LIVE = on-site performance/test (usually a blocker) · Online / None = fine.",
 "  Est. YOUR cost/yr   Rough: tuition + living - scholarship. 0 = funding covers it.",
 "                      !! Living costs are MY estimate, not from official pages. Ballpark only.",
 "  Verified            Whether a second pass re-confirmed the facts on the official page.",
 "  Link                Clickable — always check the official page before applying.",
 "", "THE NUMBERS",
 f"  Programmes found                {len(rows)}",
 f"  A-rated (your real targets)     {nA}",
 f"  B-rated / C-rated               {nB} / {nC}",
 f"  Fully funded                    {nfull}",
 f"  Accept a non-music bachelor     {nnm}",
 "", "REMEMBER",
 "  IELTS by ~Sept 2026, and 4-5 finished portfolio pieces by Nov 2026.",
 "  One portfolio covers ReSound (3 works/15min), Aalto (max 3), Glasgow (2-3), Westminster.",
]
for i, line in enumerate(guide, start=2):
    c = ws.cell(row=i, column=1, value=line)
    if line.isupper() and line.strip(): c.font = Font(bold=True, size=11, color=NAVY)
ws.column_dimensions["A"].width = 104
ws.sheet_view.showGridLines = False

sheet(wb, "A-LIST", [r for r in rows if r["Rating"]=="A"], MAIN, "2E7D32")
sheet(wb, "All Programmes", rows, MAIN, "1F4E79")
sheet(wb, "No Audition", [r for r in rows if r["Audition"] in ("No","None","Online")], MAIN, "6A1B9A")
sheet(wb, "Accepts Non-Music", [r for r in rows if r["Accepts non-music?"]=="Yes"], MAIN, "00838F")
sheet(wb, "Fully Funded", [r for r in rows if r["Funding"]=="Full"], MAIN, "EF6C00")

ws = wb.create_sheet("By Country"); ws.sheet_properties.tabColor = "455A64"
hdr = ["Country","Programmes","A-rated","Fully funded","Accept non-music","No live audition"]
ws.append(hdr)
for i, c in enumerate(hdr, 1):
    cell = ws.cell(row=1, column=i); cell.font = hf; cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30
agg = defaultdict(lambda: dict(n=0,a=0,f=0,nm=0,na=0))
for r in rows:
    g = agg[r["Country"]]; g["n"] += 1
    if r["Rating"]=="A": g["a"] += 1
    if r["Funding"]=="Full": g["f"] += 1
    if r["Accepts non-music?"]=="Yes": g["nm"] += 1
    if r["Audition"] in ("No","None","Online"): g["na"] += 1
for c, g in sorted(agg.items(), key=lambda kv: (-kv[1]["n"], kv[0])):
    ws.append([c, g["n"], g["a"], g["f"], g["nm"], g["na"]])
for i, w in enumerate([26,13,11,13,17,16], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[get_column_letter(i)][1:]:
        cell.alignment = Alignment(horizontal="center" if i > 1 else "left", vertical="center")
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:F{ws.max_row}"

sheet(wb, "Full Detail", rows, DETAIL, "9E9E9E")
wb.save(OUT)
print(f"wrote {OUT}: {wb.sheetnames}")
print(f"rows={len(rows)} A={nA} B={nB} C={nC} full={nfull} nonmusic={nnm}")
