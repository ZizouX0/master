#!/usr/bin/env python3
"""Build the Europe gap-sweep workbook.

This dataset answers three questions the earlier sweeps did not answer consistently:
  1. Is the institution PUBLIC or PRIVATE?
  2. What language will he actually be TAUGHT in (which differs from the certificate demanded)?
  3. What does a NON-EU national pay, which is often not the headline fee?
Those three are the first columns after the name, and every view is filterable on them.
"""
import json, re, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = ".masters-search/results/europe-gapfill.json"
OUT = "results/europe-gapfill.xlsx"

HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(color="FFFFFF", bold=True, size=11)
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GREY = PatternFill("solid", fgColor="EDEDED")


def txt(s, n=400):
    s = re.sub(r"\s+", " ", str(s or "")).strip()
    if not s or s.lower() in ("n/a", "na", "none", "unknown", "-"):
        return ""
    return s if len(s) <= n else s[: n - 1] + "…"


def country(r):
    """'United Kingdom (England)' and 'United Kingdom (Wales)' are one country for filtering."""
    c = re.sub(r"\s*\([^)]*\)", "", (r.get("country") or "")).strip()
    return {"UK": "United Kingdom", "Czechia": "Czech Republic",
            "Türkiye": "Turkey", "Turkiye": "Turkey"}.get(c, c) or "Unknown"


def region(r):
    """Keep the sub-national detail (England/Wales/Scotland, a German Land) visible."""
    m = re.search(r"\(([^)]*)\)", r.get("country") or "")
    return m.group(1) if m else ""


def pubpriv(r):
    s = (r.get("publicPrivate") or "").strip().lower()
    if s.startswith("public") or re.match(r"^\s*(state|government)", s):
        return "PUBLIC"
    if s.startswith("private"):
        return "PRIVATE"
    if "public" in s[:40]:
        return "PUBLIC"
    if "private" in s[:40]:
        return "PRIVATE"
    return "unclear"


LANGS = ["English", "German", "French", "Italian", "Spanish", "Portuguese", "Dutch", "Danish",
         "Swedish", "Norwegian", "Finnish", "Polish", "Czech", "Slovak", "Hungarian", "Romanian",
         "Bulgarian", "Greek", "Turkish", "Estonian", "Latvian", "Lithuanian", "Slovenian",
         "Croatian", "Serbian", "Icelandic"]


def language(r):
    """Short, filterable teaching language.

    The agents state the real teaching language first and then qualify it in prose
    ("GERMAN. The official page notes some English-language materials…"). Reading the
    whole field therefore over-reports English badly, so read the opening statement and
    only fall back to the full text when that is inconclusive.
    """
    s = (r.get("language") or "").strip()
    if not s:
        return "Check"
    head = re.split(r"[.;(]", s)[0]                       # the declarative bit
    hits = [L for L in LANGS if re.search(rf"\b{L}\b", head, re.I)]
    if not hits:
        hits = [L for L in LANGS if re.search(rf"\b{L}\b", s[:90], re.I)]
    if not hits:
        return "Check"
    if len(hits) == 1:
        return hits[0]
    # a genuine dual-language programme — name both, English first if present
    hits.sort(key=lambda L: (L != "English", L))
    return " + ".join(hits[:2])


FX = {"EUR": 1.0, "€": 1.0, "USD": 0.92, "$": 0.92, "GBP": 1.17, "£": 1.17, "DKK": 0.134,
      "SEK": 0.088, "NOK": 0.086, "CHF": 1.05, "PLN": 0.23, "HUF": 0.0026, "CZK": 0.040,
      "RON": 0.20, "BGN": 0.51, "ISK": 0.0067, "TRY": 0.026, "TL": 0.026}
CUR = "|".join(re.escape(k) for k in sorted(FX, key=len, reverse=True))
# a figure only counts as money when a currency sits next to it, on either side
RX_MONEY = re.compile(rf"(?:({CUR})\s?([\d][\d ,.]*\d)|([\d][\d ,.]*\d)\s?({CUR}))", re.I)


def _eur(raw, cur):
    n = raw.strip().replace(" ", "")
    if re.match(r"^\d{1,3}(,\d{3})+(\.\d+)?$", n):      # 264,000.00
        n = n.replace(",", "")
    elif re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", n):    # 264.000,00
        n = n.replace(".", "").replace(",", ".")
    else:
        n = n.replace(",", ".")
    try:
        v = float(n)
    except ValueError:
        return None
    if 1990 <= v <= 2100 and v == int(v):               # a year, not a price
        return None
    return v * FX.get(cur.upper(), FX.get(cur, 1.0))


# "no tuition", "no non-EU tuition", "charges no tuition", "ZERO", "does not levy…"
RX_FREE = re.compile(r"\bno (?:\S+ ){0,2}tuition|tuition[- ]free|free of charge|"
                     r"\bzero\b|eur 0\b|does not (?:charge|levy)|no fees?\b", re.I)


def money_flag(r):
    """Green when he pays little or nothing, red when it is expensive.

    Only the NON-EU field counts — that is his rate. The free-phrase test is scoped to the
    opening of that field, because a record can say "EUR 10,000 for him" and then explain
    that EU students pay no tuition; scanning the whole string flips those to FREE wrongly.
    """
    s = (r.get("tuitionNonEU") or "").strip()
    if not s:
        s = (r.get("tuitionEU") or "").strip()
    if RX_FREE.search(s[:120]):
        return "FREE"
    vals = []
    for m in RX_MONEY.finditer(s):
        cur = m.group(1) or m.group(4)
        raw = m.group(2) or m.group(3)
        v = _eur(raw, cur)
        if v and 20 <= v <= 500_000:
            vals.append(v)
    if not vals:
        return "Check"
    big = max(vals)
    if big < 1500:
        return "very cheap"
    if big < 5000:
        return "cheap"
    if big < 15000:
        return "moderate"
    return "expensive"


def yesno(v):
    s = (v or "").strip().lower()
    if s.startswith("yes"):
        return "Yes"
    if s.startswith("no") and not s.startswith("n/a"):
        return "No"
    return "Check"


def portfolio(r):
    s = (r.get("portfolioSpec") or "").strip().lower()
    if not s or s.startswith("none") or "not required" in s or "no portfolio" in s:
        return "None"
    return "Required"


def audition(r):
    s = (r.get("auditionSpec") or "").strip().lower()
    if not s or s.startswith("none") or "no audition" in s:
        return "None"
    if "interview" in s and "audition" not in s:
        return "Interview"
    return "Audition"


COLS = [
    ("Country", lambda r: country(r), 16),
    ("Region", lambda r: region(r), 14),
    ("City", lambda r: txt(r.get("city"), 30), 20),
    ("Public/Private", lambda r: pubpriv(r), 13),
    ("Taught in", lambda r: language(r), 16),
    ("Cost band", lambda r: money_flag(r), 12),
    ("Portfolio", lambda r: portfolio(r), 11),
    ("Audition", lambda r: audition(r), 11),
    ("Accepts non-music", lambda r: yesno(r.get("acceptsNonMusic")), 13),
    ("Institution", lambda r: txt(r.get("institution"), 90), 42),
    ("Programme", lambda r: txt(r.get("program"), 90), 44),
    ("YOUR tuition (non-EU)", lambda r: txt(r.get("tuitionNonEU")), 34),
    ("EU tuition", lambda r: txt(r.get("tuitionEU")), 24),
    ("Other fees", lambda r: txt(r.get("otherFees")), 30),
    ("All-in estimate", lambda r: txt(r.get("totalCostEstimate")), 34),
    ("Scholarships", lambda r: txt(r.get("scholarships")), 34),
    ("Funding level", lambda r: txt(r.get("fundingLevel"), 20), 14),
    ("Language detail", lambda r: txt(r.get("language")), 34),
    ("Language certificate", lambda r: txt(r.get("languageRequirement")), 30),
    ("English test", lambda r: txt(r.get("englishTest")), 26),
    ("Deadline", lambda r: txt(r.get("deadline")), 30),
    ("Applications open", lambda r: txt(r.get("applicationOpens")), 26),
    ("Intakes", lambda r: txt(r.get("intakes"), 60), 18),
    ("Length / ECTS", lambda r: txt(r.get("durationEcts"), 60), 18),
    ("Qualification", lambda r: txt(r.get("qualification")), 36),
    ("Accreditation", lambda r: txt(r.get("accreditationStatus")), 34),
    ("Visa OK", lambda r: yesno(r.get("visaEligible")), 9),
    ("Entry requirements", lambda r: txt(r.get("entryRequirements")), 40),
    ("Portfolio detail", lambda r: txt(r.get("portfolioSpec")), 36),
    ("Focus", lambda r: txt(r.get("focus")), 34),
    ("What you study", lambda r: txt(r.get("whatYouStudy")), 40),
    ("Facilities", lambda r: txt(r.get("facilities")), 34),
    ("Verdict", lambda r: txt(r.get("valueVerdict"), 20) or "—", 13),
    ("Rating", lambda r: txt(r.get("rating"), 4) or "—", 8),
    ("Recommendation", lambda r: txt(r.get("recommendation"), 600), 54),
    ("Source", lambda r: r.get("sourceUrl") or "", 28),
]

COST_ORDER = {"FREE": 0, "very cheap": 1, "cheap": 2, "moderate": 3, "expensive": 4, "Check": 5}


def sheet(wb, title, rows, note=None):
    ws = wb.create_sheet(title[:31])
    top = 1
    if note:
        c = ws.cell(1, 1, note)
        c.font = Font(bold=True, size=11, color="1F3864")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        top = 3
    for i, (label, _, w) in enumerate(COLS, 1):
        c = ws.cell(top, i, label)
        c.fill, c.font = HDR, HDRF
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[top].height = 30
    for n, r in enumerate(rows):
        rr = top + 1 + n
        for i, (label, fn, _) in enumerate(COLS, 1):
            ws.cell(rr, i, fn(r)).alignment = Alignment(vertical="top", wrap_text=False)
        ws.cell(rr, 4).fill = GREEN if pubpriv(r) == "PUBLIC" else (AMBER if pubpriv(r) == "PRIVATE" else GREY)
        lang = language(r)
        ws.cell(rr, 5).fill = GREEN if lang.startswith("English") or lang == "French" else GREY
        band = money_flag(r)
        ws.cell(rr, 6).fill = {"FREE": GREEN, "very cheap": GREEN, "cheap": GREEN,
                               "moderate": AMBER, "expensive": RED}.get(band, GREY)
        ws.cell(rr, 7).fill = GREEN if portfolio(r) == "None" else AMBER
        ws.cell(rr, 8).fill = GREEN if audition(r) == "None" else AMBER
        src = ws.cell(rr, len(COLS))
        u = r.get("sourceUrl") or ""
        if u.startswith("http"):
            src.hyperlink, src.value, src.font = u, "official page", Font(color="0563C1", underline="single")
    ws.auto_filter.ref = f"A{top}:{get_column_letter(len(COLS))}{top + len(rows)}"
    ws.freeze_panes = ws.cell(top + 1, 10)
    return ws


def main():
    data = json.load(open(SRC, encoding="utf-8"))
    data.sort(key=lambda r: (country(r), COST_ORDER.get(money_flag(r), 5),
                             r.get("institution") or ""))
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 108
    for i, (t, b) in enumerate([
        ("EUROPE GAP SWEEP — institutions the earlier searches missed", True),
        ("", False),
        ("Every row here is NEW. The 905 institutions already found in the earlier sweeps were", False),
        ("given to the agents as an exclusion list, so nothing is repeated from those files.", False),
        ("", False),
        ("The four columns that matter are colour-coded:", True),
        ("  Public/Private  green = public (usually far cheaper, and state-accredited by default)", False),
        ("  Taught in       green = English or French — the two languages you can already work in", False),
        ("  Cost band       green = free or under ~5,000/yr for YOU as a non-EU national", False),
        ("  Portfolio /     green = not required. These are the programmes you could apply to", False),
        ("  Audition        before your portfolio exists.", False),
        ("", False),
        ("'YOUR tuition (non-EU)' is the rate a Tunisian national pays. It is often different from", False),
        ("the headline figure, and in much of Germany and Scandinavia the difference is the whole", False),
        ("story — several German states charge non-EU students no tuition at all.", False),
        ("", False),
        ("'Taught in' vs 'Language certificate' are deliberately separate columns. A programme can", False),
        ("be taught in English and still demand a German certificate, or vice versa.", False),
        ("", False),
        ("Tabs: No Portfolio Needed · English-taught · Free or Cheap · By Country · Everything", False),
    ], 1):
        c = ws.cell(i, 1, t)
        if b:
            c.font = Font(bold=True, size=12, color="1F3864")

    noport = [r for r in data if portfolio(r) == "None" and audition(r) == "None"]
    eng = [r for r in data if language(r).startswith("English")]
    cheap = [r for r in data if money_flag(r) in ("FREE", "very cheap", "cheap")]

    if noport:
        sheet(wb, "No Portfolio Needed", noport,
              "No portfolio AND no audition — you could apply to these today.")
    if eng:
        sheet(wb, "English-taught", eng, "Taught in English.")
    if cheap:
        sheet(wb, "Free or Cheap", cheap, "Free, or under roughly 5,000 a year for a non-EU national.")
    sheet(wb, "Everything", data, "Every new institution found. Use the header arrows to filter.")

    by = defaultdict(list)
    for r in data:
        by[country(r)].append(r)
    for c in sorted(by):
        sheet(wb, c, by[c], f"{c} — {len(by[c])} new programmes.")

    wb.save(OUT)
    print(f"wrote {OUT}: {len(data)} rows | {len(noport)} no-portfolio | {len(eng)} English | "
          f"{len(cheap)} cheap | {len(by)} countries")


if __name__ == "__main__":
    main()
