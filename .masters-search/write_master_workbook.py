#!/usr/bin/env python3
"""Write the master workbook and CSV from the consolidated dataset."""
import csv, json, re, sys
from collections import defaultdict, Counter
sys.path.insert(0, '.masters-search')
import build_master as M
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(color="FFFFFF", bold=True, size=11)
GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
GREY = PatternFill("solid", fgColor="EDEDED")


def g(r, *keys, n=400):
    for k in keys:
        v = M.txt(r["_rec"].get(k), n)
        if v:
            return v
    return ""


# Identity first (these three stay frozen while you scroll right), then the six
# decision columns, then everything else.
COLS = [
    ("Country", lambda r: r["Country"], 17),
    ("Institution", lambda r: r["Institution"], 40),
    ("Programme / scheme", lambda r: r["Programme / scheme"], 42),
    ("Chance for you", lambda r: r["Chance for you"], 14),
    ("Cost band (you)", lambda r: r["Cost band (you)"], 15),
    ("Taught in", lambda r: r["Taught in"], 16),
    ("Public/Private", lambda r: r["Public/Private"], 13),
    ("Funding", lambda r: r["Funding"], 10),
    ("Why that chance", lambda r: r["Why that chance"], 46),
    ("Region", lambda r: r["Region"], 22),
    ("City", lambda r: r["City"], 20),
    ("Scholarship", lambda r: r["Scholarship"], 30),
    ("Track", lambda r: r["Track"], 26),
    ("Degree awarded", lambda r: g(r, "qualification", "degree"), 34),
    ("Accreditation", lambda r: g(r, "accreditationStatus"), 32),
    ("Visa OK", lambda r: g(r, "visaEligible", n=120), 12),
    ("YOUR tuition", lambda r: g(r, "tuitionNonEU", "tuitionIntl", "tuitionTotal", "tuitionPerYear"), 34),
    ("Other fees", lambda r: g(r, "otherFees"), 28),
    ("All-in estimate", lambda r: g(r, "totalCostEstimate"), 32),
    ("Scholarship name", lambda r: g(r, "scholarshipName"), 30),
    ("Scholarship detail", lambda r: g(r, "scholarships", "scholarship", "stipendCoverage"), 36),
    ("Language certificate", lambda r: g(r, "languageRequirement"), 28),
    ("English test", lambda r: g(r, "englishTest"), 26),
    ("Deadline", lambda r: g(r, "deadline"), 30),
    ("Applications open", lambda r: g(r, "applicationOpens"), 26),
    ("Length / ECTS", lambda r: g(r, "durationEcts", n=60), 18),
    ("Entry requirements", lambda r: g(r, "entryRequirements"), 42),
    ("Accepts non-music", lambda r: g(r, "acceptsNonMusic", n=140), 22),
    ("Portfolio", lambda r: g(r, "portfolioSpec", "portfolioReq"), 34),
    ("Audition", lambda r: g(r, "auditionSpec"), 26),
    ("What you study", lambda r: g(r, "whatYouStudy", "focus"), 40),
    ("Verdict", lambda r: g(r, "valueVerdict", n=20) or "—", 12),
    ("Rating", lambda r: g(r, "rating", n=4) or "—", 8),
    ("Recommendation", lambda r: g(r, "recommendation", "fitNotes", n=700), 56),
    ("Found by", lambda r: r["_rec"].get("_source", ""), 20),
    ("Source", lambda r: r["_rec"].get("sourceUrl") or "", 26),
]

IDX = {name: i for i, (name, _fn, _w) in enumerate(COLS, 1)}


def sheet(wb, title, rows, note=None):
    ws = wb.create_sheet(title[:31])
    top = 1
    if note:
        c = ws.cell(1, 1, note)
        c.font = Font(bold=True, size=11, color="1F3864")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        top = 3
    for i, (label, _, w) in enumerate(COLS, 1):
        c = ws.cell(top, i, label)
        c.fill, c.font = HDR, HDRF
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[top].height = 30
    for n, r in enumerate(rows):
        rr = top + 1 + n
        for i, (_, fn, _w) in enumerate(COLS, 1):
            ws.cell(rr, i, fn(r)).alignment = Alignment(vertical="top", wrap_text=False)
        ws.cell(rr, IDX["Chance for you"]).fill = {
            "Strong": GREEN, "Possible": AMBER, "Weak": RED}[r["Chance for you"]]
        ws.cell(rr, IDX["Chance for you"]).font = Font(bold=True)
        ws.cell(rr, IDX["Cost band (you)"]).fill = {
            "Free": GREEN, "Under €1,500": GREEN, "€1,500–5,000": GREEN,
            "€5,000–15,000": AMBER, "Over €15,000": RED}.get(r["Cost band (you)"], GREY)
        ws.cell(rr, IDX["Taught in"]).fill = GREEN if M.language_ok(r["Taught in"]) else GREY
        ws.cell(rr, IDX["Public/Private"]).fill = GREEN if r["Public/Private"] == "Public" else (
            AMBER if r["Public/Private"] == "Private" else GREY)
        ws.cell(rr, IDX["Funding"]).fill = {"Full": GREEN, "Partial": AMBER}.get(r["Funding"], GREY)
        src = ws.cell(rr, len(COLS))
        u = r["_rec"].get("sourceUrl") or ""
        if u.startswith("http"):
            src.hyperlink, src.value, src.font = u, "official page", Font(color="0563C1", underline="single")
    ws.auto_filter.ref = f"A{top}:{get_column_letter(len(COLS))}{top + len(rows)}"
    ws.freeze_panes = ws.cell(top + 1, 4)   # header + Country/Institution/Programme
    return ws


def main():
    rows = M.main()
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("START HERE")
    ws.column_dimensions["A"].width = 112
    lines = [
        ("EVERYTHING — every master's programme and scholarship found, in one place", True),
        ("", False),
        (f"{len(rows)} unique opportunities · {len({r['Country'] for r in rows})} countries · "
         f"merged and de-duplicated from five separate searches.", False),
        ("", False),
        ("COLUMNS A-C identify the programme and stay frozen while you scroll right.", True),
        ("COLUMNS D-H are the decision. Everything after them is supporting detail.", True),
        ("", False),
        ("  Chance for you   GREEN  Strong   — the entry rules name your background, or accept any degree", False),
        ("                   AMBER  Possible — plausible, but something needs asking or arguing", False),
        ("                   RED    Weak     — a music degree, an audition, pro experience, or a language you lack", False),
        ("  Why that chance  the actual reason, so you can judge it yourself rather than trust the label", False),
        ("  Cost band        what YOU pay as a Tunisian national — not the headline EU rate", False),
        ("  Taught in        green = English or French, the two languages you already work in", False),
        ("  Public/Private   green = public. Usually cheaper and accredited by default", False),
        ("  Funding          green = fully funded", False),
        ("", False),
        ("HOW TO USE IT: open the 'Everything' tab, filter Chance = Strong, then filter Cost band.", False),
        ("That is your real shortlist. The ready-made tabs below do the common combinations for you.", False),
        ("", False),
        ("A WARNING ABOUT 'Chance'. It is computed from the published entry text, not from an", True),
        ("admissions officer. It is a filter to save you time, not a prediction. Always read the", False),
        ("'Entry requirements' and 'Why that chance' columns before ruling anything in or out.", False),
        ("", False),
        ("Blank cells mean the institution publishes nothing — not that the answer is zero.", False),
        ("Dates marked PRIOR CYCLE are last year's, kept because these calendars repeat closely.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        c = ws.cell(i, 1, t)
        if b:
            c.font = Font(bold=True, size=12, color="1F3864")

    strong = [r for r in rows if r["Chance for you"] == "Strong"]
    cheap = [r for r in rows if r["Cost band (you)"] in ("Free", "Under €1,500", "€1,500–5,000")]
    best = [r for r in strong if r["Cost band (you)"] in ("Free", "Under €1,500", "€1,500–5,000")]
    workable = [r for r in strong if M.language_ok(r["Taught in"])]
    funded = [r for r in rows if r["Funding"] == "Full"]
    schol = [r for r in rows if r["Scholarship"] != "—"]

    sheet(wb, "★ BEST BETS", best,
          "Strong chance AND cheap or free. If you read one tab, read this one.")
    sheet(wb, "Strong + English or French", workable,
          "Strong chance, taught in a language you already work in.")
    sheet(wb, "Fully funded", funded, "Funding covers tuition and usually a stipend.")
    sheet(wb, "Has a scholarship", schol,
          "A named funding scheme is attached. Scholarships here are always tied to a programme.")
    sheet(wb, "Free or cheap", cheap, "Free, or under EUR 5,000 a year for you.")
    sheet(wb, "Everything", rows, "All records. Filter with the arrows on the header row.")

    # scholarship index — one row per scheme, so he can see what funding exists at all
    idx = defaultdict(list)
    for r in schol:
        idx[r["Scholarship"]].append(r)
    ws = wb.create_sheet("Scholarship index")
    for i, (h, w) in enumerate([("Scholarship scheme", 62), ("Programmes", 12),
                                ("Countries", 40), ("Best chance offered", 18)], 1):
        c = ws.cell(1, i, h); c.fill, c.font = HDR, HDRF
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    order = {"Strong": 0, "Possible": 1, "Weak": 2}
    for n, (name, rs) in enumerate(sorted(idx.items(), key=lambda kv: -len(kv[1])), 2):
        best_ch = sorted((x["Chance for you"] for x in rs), key=lambda c: order[c])[0]
        ws.cell(n, 1, name)
        ws.cell(n, 2, len(rs))
        ws.cell(n, 3, ", ".join(sorted({x["Country"] for x in rs}))[:250])
        c = ws.cell(n, 4, best_ch)
        c.fill = {"Strong": GREEN, "Possible": AMBER, "Weak": RED}[best_ch]
    ws.auto_filter.ref = f"A1:D{1 + len(idx)}"
    ws.freeze_panes = ws.cell(2, 1)

    by = defaultdict(list)
    for r in rows:
        by[r["Region"]].append(r)
    for reg in sorted(by, key=lambda k: -len(by[k])):
        sheet(wb, reg, by[reg], f"{reg} — {len(by[reg])} opportunities.")

    wb.save(M.OUT_XLSX)

    # Excel only — the CSV duplicate was dropped at the user's request, and the
    # workbook carries the filters and colour coding that make the data usable.
    print(f"wrote {M.OUT_XLSX}")
    print(f"  {len(rows)} rows · {len(best)} best bets · {len(workable)} strong+EN/FR · "
          f"{len(funded)} fully funded · {len(schol)} with a scheme · {len(idx)} distinct schemes · {len(by)} regions")


if __name__ == "__main__":
    main()
