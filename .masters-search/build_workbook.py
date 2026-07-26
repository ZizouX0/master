#!/usr/bin/env python3
"""Build a readable, filterable Excel workbook from opportunities.json."""
import json, re, sys
sys.path.insert(0, '.masters-search')
import cost_model
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

SRC = "results/opportunities.json"
OUT = "results/opportunities.xlsx"

# ---------- normalisation ----------
COUNTRY_FIX = {
    "usa": "USA", "united states": "USA", "us": "USA",
    "uk": "United Kingdom", "united kingdom": "United Kingdom",
    "turkiye": "Turkey", "türkiye": "Turkey",
    "czechia": "Czech Republic", "the netherlands": "Netherlands",
    "eu-wide": "Multi-country", "eu": "Multi-country",
}
REGION = {
    "Europe": {"United Kingdom","France","Germany","Netherlands","Spain","Portugal","Italy","Austria",
        "Switzerland","Denmark","Sweden","Norway","Finland","Iceland","Ireland","Belgium","Luxembourg",
        "Poland","Czech Republic","Slovakia","Hungary","Slovenia","Croatia","Serbia","Romania","Bulgaria",
        "Greece","Cyprus","Malta","Estonia","Latvia","Lithuania","Albania","North Macedonia","Bosnia",
        "Montenegro","Kosovo","Liechtenstein","Turkey","Moldova","Ukraine"},
    "North America": {"USA","Canada"},
    "Asia-Pacific": {"Australia","New Zealand","Singapore","South Korea","Japan","Hong Kong","China","Taiwan"},
    "Middle East & Africa": {"UAE","Qatar","Egypt","Tunisia","Morocco","Saudi Arabia","Israel"},
}
REGION_OF = {c: r for r, cs in REGION.items() for c in cs}

def clean_country(raw):
    """Primary country from a messy string like 'Ireland (+ Portugal, Lithuania mobility)'."""
    s = (raw or "").strip()
    if not s:
        return "Unknown"
    low = s.lower()
    if low.startswith("multiple") or "member" in low or low.startswith("eu-wide"):
        return "Multi-country"
    s = re.sub(r"\([^)]*\)", " ", s)          # drop parentheticals
    s = re.split(r"[/,;]| \+ | and ", s)[0]    # first listed country
    s = re.sub(r"\s+", " ", s).strip(" -+")
    return COUNTRY_FIX.get(s.lower(), s) or "Unknown"

def multi_country(raw):
    """Yes if the program spans several countries (Erasmus Mundus style)."""
    s = raw or ""
    return "Yes" if (re.search(r"[/,]|\+", re.sub(r"\([^)]*\)", "", s)) or "mobility" in s.lower()) else "No"

def clean_funding(raw):
    s = (raw or "").strip().lower()
    if not s: return "Check"
    if s.startswith("full"): return "Full"
    if s.startswith("partial"):
        return "Partial"
    if "full" in s:  return "Full (external scheme)"
    if s.startswith("none"): return "None"
    if s.startswith("unverified") or s.startswith("check"): return "Check"
    return "Check"

TRACK_MAP = [
    ("sound-design", "Sound Design / Production"), ("sound design", "Sound Design / Production"),
    ("sound/music", "Sound & Music Tech"), ("music-tech", "Sound & Music Tech"),
    ("music-business", "Music Business"), ("music business", "Music Business"),
    ("acoustics", "Acoustics / Audio Eng"),
    ("mba", "Business / MBA"), ("business", "Business / MBA"),
    ("computing", "Sound & Music Tech"), ("media", "Sound & Music Tech"),
]
def clean_track(raw):
    s = (raw or "").lower()
    for k, v in TRACK_MAP:
        if k in s: return v
    return "Other"

def yes_no(raw, default="Check"):
    s = (raw or "").strip().lower()
    if not s: return default
    if s.startswith("y"): return "Yes"
    if s.startswith("n/a") or s.startswith("n /"): return "Check"
    if s.startswith("n"): return "No"
    if s.startswith("check") or "unverified" in s: return "Check"
    return "Check"

def clean_language(raw):
    s = (raw or "").lower()
    if not s: return "Check"
    if "english" in s: return "English"
    if "french" in s or "français" in s: return "French"
    return (raw or "Check").split("(")[0].strip()[:20] or "Check"

def deadline_status(raw):
    s = (raw or "").lower()
    if "prior cycle" in s: return "Confirm 2027 date"
    if not s or "unverified" in s or "check" in s: return "Not published"
    return "Date published"

# ---------- admission difficulty (for THIS candidate) ----------
# Candidate: software-engineering BSc, no music degree, no professional audio experience,
# no performance training, IELTS pending, portfolio in progress; Arabic/French/English.
RX_CS       = re.compile(r"computational science|computer science|software engineer", re.I)
RX_ANY      = re.compile(r"any field|any discipline|any academic|all academic disciplines|regardless of (the )?field", re.I)
RX_OPEN     = re.compile(r"other (fields|degrees|subjects)|related field|or related|non-music|"
                         r"relevant (recognised )?bachelor|professional experience at comparable|exceptional talent", re.I)
# a bachelor's with NO subject restriction named = effectively open
RX_GENERIC  = re.compile(r"^\s*(a )?bachelor'?s?\b(?![^.;]{0,40}\b(of|in)\s+(music|art|law|econom))", re.I)
RX_MUSICBA  = re.compile(r"bachelor of music|bachelor'?s? (degree )?in music|instrumental bachelor|"
                         r"music bachelor|ba in music", re.I)
RX_PERF     = re.compile(r"live (performance|computer-performance) demonstration|in-person (entrance|audition)|audition", re.I)
RX_ONLINE   = re.compile(r"online audition|video (recording|audition)|no live admission|recorded audition|"
                         r"without a live audition", re.I)
RX_PROEXP   = re.compile(r"significant experience[^.]{0,60}professional level|professional[- ]level "
                         r"(experience|audio)|\d[,.]?\d*00 hours|years of (relevant )?work experience|"
                         r"experience is mandatory|professional experience[^.]{0,30}required", re.I)
RX_SKILLTEST= re.compile(r"on-the-spot|mix test|stereo mix of a pre-r", re.I)
RX_GPA      = re.compile(r"gpa\s*>?=?\s*[23]\.\d|2:1|second class", re.I)
RX_AGE      = re.compile(r"under 30|age limit|born on or after", re.I)
RX_PREFER   = re.compile(r"preferably in (a )?related|art history|cultural management, economics, law", re.I)
LANG_BLOCK  = ("spanish","portuguese","german","italian","greek","korean","romanian","bulgarian",
               "serbian","polish","czech","hungarian","dutch","danish","swedish","norwegian",
               "finnish","turkish","japanese","chinese","slovak","slovenian","croatian",
               "lithuanian","latvian","estonian")
RX_LANG_UNSURE = re.compile(r"conflict|check|depend|n/?a|programme-dependent|host", re.I)
TECH_TRACKS = ("Sound & Music Tech", "Acoustics / Audio Eng")

def admission_difficulty(o, lang, elig, track):
    """Green / Amber / Red + reason, judged from the official entry text."""
    txt = " ".join(str(o.get(k, "")) for k in ("entryRequirements", "portfolioReq", "program", "fitNotes"))
    entry = str(o.get("entryRequirements", ""))
    low = (lang or "").lower()
    lang_unsure = bool(RX_LANG_UNSURE.search(low)) or low in ("", "check")

    # ---- hard blockers ----
    if elig == "No":
        return "Red", "Tunisian nationals not eligible"
    if low.startswith(LANG_BLOCK) and not lang_unsure:
        return "Red", f"Taught in {lang} — you don't have the language"
    if RX_PROEXP.search(txt):
        return "Red", "Requires professional-level audio/work experience you don't have yet"
    if RX_MUSICBA.search(txt) and not (RX_OPEN.search(txt) or RX_CS.search(txt) or RX_ANY.search(txt)):
        return "Red", "Requires a Bachelor of Music — you don't have one"
    if RX_PERF.search(txt) and not RX_ONLINE.search(txt) and not (RX_CS.search(txt) or RX_ANY.search(txt)):
        return "Red", "Live audition — assumes performance training"

    # ---- your background explicitly fits ----
    if RX_CS.search(txt):
        return "Green", "Officially accepts computer/computational science degrees"
    if RX_ANY.search(txt):
        return "Green", "Accepts a bachelor's in any field"
    if RX_SKILLTEST.search(txt):
        return "Amber", "Practical skills test on the day — prepare hard"
    if RX_GENERIC.search(entry.strip()) and not RX_MUSICBA.search(txt):
        return "Green", "Open entry — a bachelor's in any subject, no field restriction named"
    if track in TECH_TRACKS and not RX_MUSICBA.search(txt) and not RX_PERF.search(txt):
        return "Green", "Technical program — an engineering background is the expected profile"
    if RX_PREFER.search(txt):
        return "Amber", "Prefers arts/economics/law backgrounds — you'd argue your case"
    if RX_OPEN.search(txt):
        return "Green", "Open to other/related fields with relevant experience"

    # ---- clearable gates ----
    if elig == "Check":
        return "Amber", "Tunisian eligibility not confirmed — verify first"
    if lang_unsure:
        return "Amber", "Language of instruction unclear — verify before applying"
    if RX_GPA.search(txt):
        return "Amber", "GPA threshold — check against your transcript"
    if RX_AGE.search(txt):
        return "Amber", "Age limit applies — confirm you qualify"
    if RX_PERF.search(txt):
        return "Amber", "Audition required, but a recorded/online route exists"
    return "Amber", "No explicit statement on non-music backgrounds — ask admissions"

def fit_score(o, fund, track, ver):
    """Rough heuristic to sort by overall attractiveness for this candidate."""
    s = 0
    s += {"Full": 40, "Full (external scheme)": 30, "Partial": 18, "Check": 5, "None": 0}.get(fund, 0)
    s += {"Sound Design / Production": 25, "Sound & Music Tech": 25,
          "Music Business": 15, "Acoustics / Audio Eng": 13, "Business / MBA": 8}.get(track, 5)
    s += {"Verified": 15, "Partial": 8}.get(ver, 0)
    if yes_no(o.get("tunisianEligible")) == "Yes": s += 10
    if clean_language(o.get("language")) == "English": s += 10
    return s

# ---------- load + enrich ----------
data = json.load(open(SRC, encoding="utf-8"))
rows = []
for o in data:
    fund = clean_funding(o.get("fundingLevel"))
    track = clean_track(o.get("track"))
    ver = (o.get("verified") or "Unverified").title()
    country = clean_country(o.get("country"))
    lang_c = clean_language(o.get("language"))
    elig_c = yes_no(o.get("tunisianEligible"))
    diff, diff_why = admission_difficulty(o, lang_c, elig_c, track)
    rows.append({
        "Program": o.get("program", ""),
        "University": o.get("university", ""),
        "Country": country,
        "Region": REGION_OF.get(country, "Other"),
        "Multi-country?": multi_country(o.get("country")),
        "Track": track,
        "Funding": fund,
        "Fit score": fit_score(o, fund, track, ver),
        "Admission difficulty": diff,
        "Why (difficulty)": diff_why,
        "Language": clean_language(o.get("language")),
        "Portfolio?": yes_no(o.get("portfolioReq")),
        "Tunisian eligible?": yes_no(o.get("tunisianEligible")),
        "Verified": ver,
        "Deadline status": deadline_status(o.get("deadline")),
        "Scholarship": o.get("scholarshipName", ""),
        "Stipend / coverage": o.get("stipendCoverage", ""),
        "Tuition (intl)": o.get("tuitionIntl", ""),
        "Application opens": o.get("applicationOpens", ""),
        "Deadline": o.get("deadline", ""),
        "Fit notes": o.get("fitNotes", ""),
        "Source URL": o.get("sourceUrl", ""),
        "Entry requirements": o.get("entryRequirements", ""),
        "Public/Private": o.get("publicPrivate", ""),
        "All countries (raw)": o.get("country", ""),
        "Funding detail (raw)": o.get("fundingLevel", ""),
        "Verification notes": o.get("verificationNotes", ""),
    })
    rows[-1].update(cost_model.estimate(rows[-1], o))
rows.sort(key=lambda r: (-r["Fit score"], r["Country"], r["Program"]))

MAIN = ["Program","University","Country","Region","Track","Funding","Fit score",
        "Admission difficulty","Why (difficulty)","Language",
        "Portfolio?","Tunisian eligible?","Verified","Deadline status","Est. YOUR cost/yr (EUR)","Cost verdict","Scholarship",
        "Stipend / coverage","Tuition (intl)","Application opens","Deadline","Fit notes","Source URL"]
DETAIL = MAIN + ["Est. living/yr (EUR)","Est. tuition/yr (EUR)",
                 "Est. year-1 cost incl. setup (EUR)","Cost confidence","Multi-country?","Entry requirements","Public/Private","All countries (raw)",
                 "Funding detail (raw)","Verification notes"]
WIDTH = {"Program":46,"University":34,"Country":16,"Region":15,"Track":24,"Funding":20,"Fit score":9,
         "Language":12,"Admission difficulty":15,"Why (difficulty)":52,"Est. YOUR cost/yr (EUR)":16,"Cost verdict":46,
         "Est. living/yr (EUR)":15,"Est. tuition/yr (EUR)":15,
         "Est. year-1 cost incl. setup (EUR)":18,"Cost confidence":13,"Portfolio?":11,"Tunisian eligible?":15,"Verified":10,"Deadline status":17,
         "Scholarship":34,"Stipend / coverage":42,"Tuition (intl)":28,"Application opens":26,
         "Deadline":34,"Fit notes":70,"Source URL":40,"Multi-country?":13,"Entry requirements":70,
         "Public/Private":14,"All countries (raw)":28,"Funding detail (raw)":36,"Verification notes":70}

NAVY = "10233F"; BLUE = "2F6DF6"
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor=NAVY)
GREEN = PatternFill("solid", fgColor="C6EFCE")
LIGHTGREEN = PatternFill("solid", fgColor="E2F0D9")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")
RED = PatternFill("solid", fgColor="FFC7CE")

def add_sheet(wb, title, records, cols=MAIN, tab_color=None):
    ws = wb.create_sheet(title)
    if tab_color: ws.sheet_properties.tabColor = tab_color
    ws.append(cols)
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for r in records:
        ws.append([r.get(c, "") for c in cols])
    # widths + alignment (wrap OFF keeps rows compact and scannable)
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTH.get(c, 20)
        for cell in ws[get_column_letter(i)][1:]:
            cell.alignment = Alignment(vertical="center", horizontal="center" if c in
                ("Fit score","Portfolio?","Tunisian eligible?","Verified","Language","Region","Admission difficulty","Est. YOUR cost/yr (EUR)","Cost confidence") else "left")
    # clickable URLs
    if "Source URL" in cols:
        ci = cols.index("Source URL") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=ci); url = str(cell.value or "")
            if url.startswith("http"):
                cell.hyperlink = url; cell.font = Font(color=BLUE, underline="single", size=10)
    # colour the Funding column
    if "Funding" in cols:
        L = get_column_letter(cols.index("Funding") + 1); rng = f"{L}2:{L}{ws.max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Full"'], fill=GREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Full (external scheme)"'], fill=LIGHTGREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Partial"'], fill=YELLOW))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"None"'], fill=GREY))
    if "Admission difficulty" in cols:
        L = get_column_letter(cols.index("Admission difficulty") + 1); rng = f"{L}2:{L}{ws.max_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Green"'], fill=GREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Amber"'], fill=YELLOW))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Red"'], fill=RED))
    ws.freeze_panes = "C2"                      # keep Program + University visible
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    return ws

wb = Workbook(); wb.remove(wb.active)

# ---- 1. Start Here ----
ws = wb.create_sheet("START HERE"); ws.sheet_properties.tabColor = BLUE
t = ws.cell(row=1, column=1, value="Master's & Scholarship Search — how to use this file")
t.font = Font(bold=True, size=16, color=NAVY)
full_n = sum(1 for r in rows if r["Funding"] == "Full")
ext_n  = sum(1 for r in rows if r["Funding"] == "Full (external scheme)")
part_n = sum(1 for r in rows if r["Funding"] == "Partial")
ver_n  = sum(1 for r in rows if r["Verified"] == "Verified")
guide = [
    "", "TABS",
    "  All Opportunities   Every program (546). Filter it any way you like.",
    "  Fully Funded        Only full scholarships — start here if funding is the priority.",
    "  BEST BETS           Fully funded AND your background qualifies. Your real application list.",
    "  Sound & Music Tech  Your primary track (music technology, sound & music computing, audio DSP).",
    "  Sound Design/Prod   Production, sound design, studio & conservatory programs.",
    "  Music Business      Music business, arts & cultural management.",
    "  By Country          Counts per country, so you can see where the options are.",
    "  Full Detail         Same rows plus entry requirements & verification notes.",
    "", "HOW TO FILTER (the arrows in row 1)",
    "  Click the small arrow on any column header, untick 'Select All', tick what you want.",
    "  Example — everything in Germany:  Country -> tick 'Germany'.",
    "  Example — funded options in Europe:  Region -> Europe,  Funding -> Full.",
    "  Example — no portfolio needed:  Portfolio? -> No.",
    "  To clear a filter: click the arrow -> 'Clear Filter'. (Data > Clear clears all.)",
    "", "COLUMNS WORTH KNOWING",
    "  Fit score            Rough 0-100 guide (funding + track match + verified + eligibility + English).",
    "                       Rows are pre-sorted by it. It is a helper, not gospel — read Fit notes.",
    "  Admission difficulty Can YOU get in?  Green = your degree qualifies.  Amber = a gate to clear",
    "                       (audition, GPA, eligibility).  Red = blocked (needs a music degree, a live",
    "                       audition, pro experience, a language you lack, or Tunisians ineligible).",
    "                       Read 'Why (difficulty)' for the reason. It is read from the official entry",
    "                       text, so treat it as a strong hint — always confirm on the source page.",
    "  Fit score vs Difficulty   Fit score = how good the PRIZE is.  Difficulty = your CHANCE.",
    "                       A 100 with a Red is a great program you cannot get into yet.",
    "  Est. YOUR cost/yr    Ballpark of what YOU pay per year = tuition + living - scholarship.",
    "                       0 means the funding covers it. 'Cost verdict' spells it out in words.",
    "                       !! Tuition/stipend are parsed from the verified official text, but the",
    "                       LIVING-COST part is my own country estimate (modest student lifestyle),",
    "                       NOT from official pages. Check 'Cost confidence'. Budget properly before",
    "                       you commit - city matters (Paris/Amsterdam/Dublin cost more than average).",
    "  Funding              Full / Full (external scheme) / Partial / None / Check  — colour coded.",
    "                       'Full (external scheme)' = program has no money itself, but Chevening/DAAD etc. can fund it.",
    "  Country              Cleaned to ONE main country so filtering works. Full list in 'All countries (raw)'.",
    "  Multi-country?       Yes = you rotate between countries (Erasmus Mundus style).",
    "  Deadline status      'Confirm 2027 date' = only last cycle's date is published yet (normal for now).",
    "  Verified             Verified / Partial / Unverified — how much was confirmed on the official page.",
    "  Source URL           Clickable — always check the official page before applying.",
    "", "THE NUMBERS",
    f"  Total opportunities            {len(rows)}",
    f"  Fully funded                   {full_n}",
    f"  Fundable via external scheme   {ext_n}",
    f"  Partially funded               {part_n}",
    f"  Verified on official pages     {ver_n}",
    "", "REMEMBER",
    "  Two things gate almost everything: take IELTS (~Sept 2026) and build your portfolio.",
    "  Earliest hard deadline: Chevening, 6 October 2026.",
]
for i, line in enumerate(guide, start=2):
    c = ws.cell(row=i, column=1, value=line)
    if line.isupper() and line.strip():
        c.font = Font(bold=True, size=11, color=NAVY)
ws.column_dimensions["A"].width = 105
ws.sheet_view.showGridLines = False

# ---- 2. main sheets ----
add_sheet(wb, "All Opportunities", rows, MAIN, "1F4E79")
add_sheet(wb, "Fully Funded",
          [r for r in rows if r["Funding"] in ("Full", "Full (external scheme)")], MAIN, "2E7D32")
add_sheet(wb, "BEST BETS",
          [r for r in rows if r["Admission difficulty"] == "Green"
           and r["Funding"].startswith("Full")], MAIN, "C62828")
add_sheet(wb, "Sound & Music Tech",
          [r for r in rows if r["Track"] == "Sound & Music Tech"], MAIN, "6A1B9A")
add_sheet(wb, "Sound Design & Prod",
          [r for r in rows if r["Track"] == "Sound Design / Production"], MAIN, "AD1457")
add_sheet(wb, "Music Business",
          [r for r in rows if r["Track"] in ("Music Business", "Business / MBA")], MAIN, "EF6C00")

# ---- 3. By Country ----
ws = wb.create_sheet("By Country"); ws.sheet_properties.tabColor = "455A64"
hdr = ["Country", "Region", "Programs", "Fully funded", "Via external scheme", "Partial", "Verified"]
ws.append(hdr)
for i, c in enumerate(hdr, 1):
    cell = ws.cell(row=1, column=i); cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30
agg = defaultdict(lambda: dict(n=0, full=0, ext=0, part=0, ver=0, region=""))
for r in rows:
    a = agg[r["Country"]]; a["n"] += 1; a["region"] = r["Region"]
    if r["Funding"] == "Full": a["full"] += 1
    if r["Funding"] == "Full (external scheme)": a["ext"] += 1
    if r["Funding"] == "Partial": a["part"] += 1
    if r["Verified"] == "Verified": a["ver"] += 1
for country, a in sorted(agg.items(), key=lambda kv: (-kv[1]["n"], kv[0])):
    ws.append([country, a["region"], a["n"], a["full"], a["ext"], a["part"], a["ver"]])
for i, w in enumerate([24, 20, 11, 14, 20, 11, 11], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[get_column_letter(i)][1:]:
        cell.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="center")
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:G{ws.max_row}"

# ---- 4. Full detail ----
add_sheet(wb, "Full Detail", rows, DETAIL, "9E9E9E")

wb.save(OUT)
print(f"wrote {OUT}: {len(wb.sheetnames)} sheets -> {wb.sheetnames}")
print(f"rows={len(rows)} full={full_n} external={ext_n} partial={part_n} verified={ver_n}")
